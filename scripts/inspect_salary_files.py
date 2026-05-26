"""
Inspect salary Excel files for years 111-115.
Mirrors the parsing logic from public/finance/index.html salaryUploadBtn handler:
- Find sheets matching \d{3,4}\d{2} pattern (民國年+月)
- Auto-detect columns by header: 姓名 / 部門 / 本薪 / 薪資小計 / 合計 / 事務所負擔勞保 / 健保 / 提撥勞退 / 事務所支出合計
- Year-end bonus sheet: f"{year}年終"
- Valid name: contains CJK/letter, not 小計/合計/外包/姓名, base_salary >= 10000
"""
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

FILES = [
    ("111", r"C:\Users\admin\Downloads\喆律111年度薪資 (1).xlsx"),
    ("112", r"C:\Users\admin\Downloads\喆律112年度薪資 (1).xlsx"),
    ("113", r"C:\Users\admin\Downloads\喆律113年度薪資 (1).xlsx"),
    ("114", r"C:\Users\admin\Downloads\（稅務第2版）喆律114年度薪資 .xlsx"),
    ("115", r"C:\Users\admin\Downloads\喆律115年薪資 (4).xlsx"),
]


def is_valid_name(n):
    if not n:
        return False
    n = str(n).strip()
    if not n:
        return False
    if re.fullmatch(r"\d[\d\s,.]*", n):
        return False
    if not re.search(r"[一-鿿㐀-䶿a-zA-Z]", n):
        return False
    if any(k in n for k in ("小計", "合計", "外包", "姓名")):
        return False
    return True


def detect_columns(rows):
    for r_idx in range(min(3, len(rows))):
        row = rows[r_idx]
        idx = {"name": -1, "dept": -1, "baseSalary": -1, "totalSalary": -1,
               "totalPay": -1, "labor": -1, "health": -1, "pension": -1, "empTotal": -1}
        for c, val in enumerate(row):
            h = (str(val) if val is not None else "").strip()
            if h == "姓名":
                idx["name"] = c
            elif h == "部門":
                idx["dept"] = c
            elif "本薪" in h:
                idx["baseSalary"] = c
            elif "薪資小計" in h:
                idx["totalSalary"] = c
            elif h.startswith("合計"):
                idx["totalPay"] = c
            elif "負擔勞保" in h:
                idx["labor"] = c
            elif "負擔健保" in h:
                idx["health"] = c
            elif "提撥勞退" in h:
                idx["pension"] = c
            elif "事務所支出合計" in h:
                idx["empTotal"] = c
        if idx["name"] >= 0 and idx["baseSalary"] >= 0:
            return idx, r_idx
    return None, -1


def inspect_file(year_label, path):
    print(f"\n{'='*78}")
    print(f"📄 {year_label}年: {Path(path).name}")
    print(f"{'='*78}")
    if not Path(path).exists():
        print(f"  ❌ 檔案不存在")
        return

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as ex:
        print(f"  ❌ 讀檔失敗: {ex}")
        return

    all_sheets = wb.sheetnames
    print(f"  全部 sheet 共 {len(all_sheets)} 個: {all_sheets[:8]}{' ...' if len(all_sheets)>8 else ''}")

    # Find month sheets
    month_sheets = []
    for n in all_sheets:
        m = re.fullmatch(r"(\d{3,4})(\d{2})", n)
        if m:
            mn = int(m.group(2))
            if 1 <= mn <= 12:
                month_sheets.append((n, int(m.group(1)), mn))
    month_sheets.sort(key=lambda x: x[2])
    print(f"  月份 sheet: {[s[0] for s in month_sheets]}  → 共 {len(month_sheets)} 個月")

    if not month_sheets:
        print(f"  ⚠️  找不到 \\d{{3,4}}\\d{{2}} 格式的月份 sheet — 解析會失敗")
        # Show first few sheet names
        print(f"  hint: 全 sheet 名稱 = {all_sheets}")
        return

    # Bonus sheet
    bonus_candidates = [f"{year_label}年終", f"{int(year_label)}年終"]
    bonus_sheet = None
    for c in bonus_candidates + [n for n in all_sheets if "年終" in n]:
        if c in all_sheets:
            bonus_sheet = c
            break
    print(f"  年終 sheet: {bonus_sheet or '(無)'}")

    # Inspect each month
    person_data = {}  # name → {months:set, dept:str}
    column_detect_log = []
    for sheet_name, yr_part, m in month_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        ci, header_row = detect_columns(rows)
        if ci is None:
            column_detect_log.append((sheet_name, "❌ fail"))
            continue
        else:
            mode = "auto"
            column_detect_log.append((sheet_name, f"✓ header at row {header_row+1}, dept_col={ci['dept']}, empTotal_col={ci['empTotal']}"))

        valid_count = 0
        no_dept_count = 0
        for i in range(header_row + 1, len(rows)):
            row = rows[i]
            if ci["name"] >= len(row): continue
            name = row[ci["name"]]
            name = str(name).strip() if name is not None else ""
            try:
                bs_raw = row[ci["baseSalary"]] if ci["baseSalary"] < len(row) else 0
                bs = float(bs_raw) if bs_raw not in (None, "") else 0
            except (ValueError, TypeError):
                bs = 0
            if not is_valid_name(name) or bs < 10000:
                continue
            valid_count += 1
            dept = ""
            if ci["dept"] >= 0 and ci["dept"] < len(row):
                d_raw = row[ci["dept"]]
                dept = str(d_raw).strip() if d_raw is not None else ""
            if not dept:
                no_dept_count += 1
            if name not in person_data:
                person_data[name] = {"months": set(), "dept": dept}
            person_data[name]["months"].add(m)
            if dept and not person_data[name]["dept"]:
                person_data[name]["dept"] = dept

        print(f"  · {sheet_name} ({m}月): 有效員工 {valid_count} 人, dept 為空 {no_dept_count} 人")

    # Column detection summary
    print(f"\n  欄位偵測：")
    for sn, msg in column_detect_log:
        print(f"    {sn}: {msg}")

    # Headcount summary
    if person_data:
        # Per-month headcount
        per_month = {}
        for n, pd_ in person_data.items():
            for m in pd_["months"]:
                per_month[m] = per_month.get(m, 0) + 1
        print(f"\n  逐月在職人數: " + ", ".join(f"{m}月={per_month.get(m,0)}" for _, _, m in month_sheets))

        total_distinct = len(person_data)
        full_year_count = sum(1 for pd_ in person_data.values() if len(pd_["months"]) == len(month_sheets))
        partial_count = total_distinct - full_year_count
        print(f"  全年共 {total_distinct} 人曾在職（完整在職 {full_year_count}，部分月份 {partial_count}）")

        # Dept coverage
        with_dept = sum(1 for pd_ in person_data.values() if pd_["dept"])
        without_dept = total_distinct - with_dept
        print(f"  有部門資料: {with_dept} 人，無部門: {without_dept} 人")

        # Distinct depts
        depts = sorted({pd_["dept"] for pd_ in person_data.values() if pd_["dept"]})
        print(f"  部門種類 ({len(depts)}): {depts}")

        # People with partial months (potential leavers/newcomers)
        partial_list = [(n, sorted(pd_["months"])) for n, pd_ in person_data.items() if len(pd_["months"]) < len(month_sheets)]
        if partial_list:
            print(f"\n  非全年在職者 ({len(partial_list)} 人):")
            for n, mlist in partial_list[:20]:
                gaps = []
                # detect gap
                full_set = set(m for _,_,m in month_sheets)
                missing = sorted(full_set - set(mlist))
                gaps_str = ",".join(str(m) for m in missing)
                print(f"    {n}: 在職 {mlist[0]}-{mlist[-1]}月（缺 {gaps_str}）部門={person_data[n]['dept']}")
            if len(partial_list) > 20:
                print(f"    ... 還有 {len(partial_list)-20} 人")


if __name__ == "__main__":
    for yr, p in FILES:
        inspect_file(yr, p)
