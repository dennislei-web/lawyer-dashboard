"""
daily_update.py
每日一鍵更新：爬取 CRM 諮詢統計 → 更新 xlsx → 更新 Supabase。

使用方式：
  python daily_update.py                     # 更新本月
  python daily_update.py --months 3          # 更新最近 3 個月
  python daily_update.py --month 2026-03     # 更新指定月份
  python daily_update.py --all               # 全部重抓 (2020 至今)

環境變數（或 .env）：
  SUPABASE_URL=https://xxxxx.supabase.co
  SUPABASE_SERVICE_KEY=eyJxxxxxxxxx
  CRM_USERNAME=your@email.com
  CRM_PASSWORD=your_password
"""

import argparse
import html as html_mod
import json
import os
import re
import sys
import time
from datetime import datetime

# Windows 終端 UTF-8 輸出
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

import pandas as pd
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv

load_dotenv()

# ─── 設定 ───────────────────────────────────────────────────
SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_KEY"]

CRM_BASE_URL = "https://crm.lawyer"
CRM_LOGIN_URL = f"{CRM_BASE_URL}/users/sign_in"
CRM_USERNAME = os.environ.get("CRM_USERNAME", "")
CRM_PASSWORD = os.environ.get("CRM_PASSWORD", "")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "consultation_all_data.xlsx")

SIGNED_MAP = {
    "initial": "未填寫",
    "unsigned": "未簽約",
    "signed_and_paid_in_full": "已簽約已付清",
    "signed_with_office_installment": "已簽約事務所分期付款",
    "signed_unpaid": "已簽約未付款",
}

CONSULT_HEADERS = [
    "諮詢日期", "案件編號", "接案所", "品牌", "諮詢律師",
    "當事人", "諮詢方式", "服務項目", "簽約狀態",
    "應收金額（案件委任金）", "已收金額（該案已收金額）", "是否列入計算",
]


# ═══════════════════════════════════════════════════════════
#  Step 1: 爬取 CRM 諮詢統計
# ═══════════════════════════════════════════════════════════

def crm_login(email, password):
    """登入 CRM，回傳 requests.Session。"""
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8",
    })

    resp = session.get(CRM_LOGIN_URL)
    soup = BeautifulSoup(resp.text, "html.parser")

    token = None
    meta = soup.find("meta", {"name": "csrf-token"})
    if meta:
        token = meta.get("content")
    if not token:
        hidden = soup.find("input", {"name": "authenticity_token"})
        if hidden:
            token = hidden.get("value")

    login_data = {
        "user[email]": email,
        "user[password]": password,
        "user[remember_me]": "1",
        "commit": "登入",
    }
    if token:
        login_data["authenticity_token"] = token

    resp = session.post(CRM_LOGIN_URL, data=login_data, allow_redirects=True)
    if "sign_in" not in resp.url and "login" not in resp.url:
        return session

    # 備用格式
    login_data_alt = {"email": email, "password": password}
    if token:
        login_data_alt["authenticity_token"] = token
    resp = session.post(CRM_LOGIN_URL, data=login_data_alt, allow_redirects=True)
    if "sign_in" not in resp.url and "login" not in resp.url:
        return session

    raise Exception("CRM 登入失敗，請確認帳號密碼")


def parse_consult_html(html_content):
    """解析 CRM 諮詢統計頁面的 React props JSON。"""
    pattern = r'data-react-class="Statistics/ConsultationStatistic/index"\s+data-react-props="([^"]*)"'
    match = re.search(pattern, html_content)
    if not match:
        return []

    decoded = html_mod.unescape(match.group(1))
    try:
        data = json.loads(decoded)
    except Exception:
        return []

    rows = []
    for r in data.get("data", []):
        try:
            dt = datetime.fromisoformat(
                r["official_date"].replace("Z", "+00:00")
            ).strftime("%Y-%m-%d")
        except Exception:
            dt = r.get("official_date", "")

        cases = r.get("cases", [])
        case = cases[0] if cases else {}
        serial = case.get("serial_number", "")
        dept = case.get("department", {}).get("name", "")
        office = case.get("council_office", {}).get("name", "")
        lawyers = ", ".join([p["name"] for p in r.get("relation_people", [])])
        clients = ", ".join(
            [cl["name"] for cl in case.get("clients", [])]
            + [cc["company_name"] for cc in case.get("client_companies", [])]
        )
        consult_type = r.get("consultation_type", {}).get("name", "")
        service_items = []
        total_accrued = total_paid = 0
        for item in case.get("case_service_items", []):
            for i in item.get("items", []):
                if i["name"] not in service_items:
                    service_items.append(i["name"])
            total_accrued += item.get("accrued_expense") or 0
            for p in item.get("payment_transactions", []):
                if not p.get("is_void", False):
                    total_paid += p.get("amount") or 0

        signed = SIGNED_MAP.get(
            r.get("signed_state", ""), r.get("signed_state", "")
        )
        is_used = "是" if r.get("is_used", False) else "否"
        rows.append([
            dt, serial, office, dept, lawyers, clients, consult_type,
            ", ".join(service_items), signed, total_accrued, total_paid, is_used,
        ])
    return rows


def scrape_month(session, year, month):
    """爬取指定月份的諮詢統計。"""
    url = f"{CRM_BASE_URL}/dashboard/statistics/consultation_statistics"
    resp = session.get(url, params={"year": str(year), "month": f"{month:02d}"})
    if resp.status_code != 200:
        return []
    return parse_consult_html(resp.text)


def read_existing_xlsx():
    """讀取現有 xlsx 資料。"""
    if not os.path.exists(XLSX_PATH):
        return []
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_list = list(row)
        if isinstance(row_list[0], datetime):
            row_list[0] = row_list[0].strftime("%Y-%m-%d")
        elif row_list[0] is not None:
            row_list[0] = str(row_list[0])
        rows.append(row_list)
    wb.close()
    return rows


def save_xlsx(all_data):
    """儲存為格式化 xlsx。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "諮詢統計"

    hf = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    hfl = PatternFill("solid", fgColor="2C3E50")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(CONSULT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf
        cell.fill = hfl
        cell.alignment = ha
        cell.border = tb

    alt = PatternFill("solid", fgColor="F2F3F4")
    df = Font(size=10, name="Arial")
    for ri, row_data in enumerate(all_data, 2):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = df
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = tb
            if ci in [10, 11]:
                cell.number_format = "#,##0"
            if ri % 2 == 0:
                cell.fill = alt

    for i, w in enumerate([12, 16, 10, 12, 20, 20, 12, 30, 20, 22, 24, 14], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(XLSX_PATH)


def scrape_crm(target_months, email, password):
    """爬取 CRM 並更新 xlsx，回傳更新後的資料筆數。"""
    print(f"\n{'='*50}")
    print("  Step 1: 爬取 CRM 諮詢統計")
    print(f"{'='*50}")

    existing = read_existing_xlsx()
    print(f"  現有 xlsx 資料：{len(existing)} 筆")

    print("  登入 CRM...", end=" ", flush=True)
    session = crm_login(email, password)
    print("OK")

    new_rows = []
    for y, m in sorted(target_months):
        print(f"    {y}/{m:02d}...", end=" ", flush=True)
        rows = scrape_month(session, y, m)
        new_rows.extend(rows)
        if rows:
            print(f"{len(rows)} 筆")
        else:
            print("無資料")
        time.sleep(0.5)

    # 合併：移除目標月份舊資料 + 加入新資料
    filtered = []
    for row in existing:
        if row[0]:
            try:
                d = datetime.strptime(str(row[0])[:10], "%Y-%m-%d")
                if (d.year, d.month) in target_months:
                    continue
            except Exception:
                pass
        filtered.append(row)

    merged = sorted(filtered + new_rows, key=lambda x: str(x[0]) if x[0] else "")
    save_xlsx(merged)
    print(f"  xlsx 已更新：{XLSX_PATH}（共 {len(merged)} 筆）")
    return len(merged)


# ═══════════════════════════════════════════════════════════
#  Step 2: 更新 Supabase
# ═══════════════════════════════════════════════════════════

def update_supabase(target_months):
    """讀取 xlsx 並更新 Supabase monthly_stats。"""
    print(f"\n{'='*50}")
    print("  Step 2: 更新 Supabase")
    print(f"{'='*50}")

    import httpx

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }

    # 讀取 xlsx
    df = pd.read_excel(XLSX_PATH)
    df.columns = df.columns.str.strip()
    print(f"  讀取 xlsx：{len(df)} 筆")
    print(f"  xlsx 欄位：{list(df.columns)}")

    # 篩選列入計算
    if "是否列入計算" in df.columns:
        df = df[df["是否列入計算"].astype(str).str.strip() != "否"].copy()

    df["諮詢日期"] = pd.to_datetime(df["諮詢日期"], errors="coerce")
    df = df.dropna(subset=["諮詢日期"])
    df["month"] = df["諮詢日期"].dt.strftime("%Y-%m")

    # 只處理目標月份
    month_strs = {f"{y}-{m:02d}" for y, m in target_months}
    df = df[df["month"].isin(month_strs)]
    if df.empty:
        print("  目標月份沒有資料")
        return 0

    # 判斷簽約
    def is_signed(status):
        s = str(status).strip()
        if s == "" or s == "nan" or "未" in s:
            return False
        return True

    df["已簽約"] = df["簽約狀態"].apply(is_signed)

    # 欄位名稱相容（清除逗號、空白等格式）
    rev_col = next((c for c in df.columns if "應收" in c), None)
    col_col = next((c for c in df.columns if "已收" in c), None)
    if rev_col:
        df["revenue"] = pd.to_numeric(
            df[rev_col].astype(str).str.replace(",", "").str.replace(" ", "").str.strip(),
            errors="coerce"
        ).fillna(0)
        print(f"  應收金額欄位: {rev_col}, 非零筆數: {(df['revenue'] > 0).sum()}, 合計: {df['revenue'].sum():,.0f}")
    else:
        print("  ⚠ 找不到應收金額欄位")
    if col_col:
        df["collected"] = pd.to_numeric(
            df[col_col].astype(str).str.replace(",", "").str.replace(" ", "").str.strip(),
            errors="coerce"
        ).fillna(0)
        print(f"  已收金額欄位: {col_col}, 非零筆數: {(df['collected'] > 0).sum()}, 合計: {df['collected'].sum():,.0f}")
    else:
        print("  ⚠ 找不到已收金額欄位")

    grouped = df.groupby(["諮詢律師", "month"]).agg(
        consult_count=("諮詢律師", "count"),
        signed_count=("已簽約", "sum"),
        revenue=("revenue", "sum") if "revenue" in df.columns else ("諮詢律師", "count"),
        collected=("collected", "sum") if "collected" in df.columns else ("諮詢律師", "count"),
    ).reset_index()
    grouped["signed_count"] = grouped["signed_count"].astype(int)
    grouped["sign_rate"] = (grouped["signed_count"] / grouped["consult_count"] * 100).round(2)

    # 取得律師 ID 對應
    with httpx.Client(timeout=30) as client:
        resp = client.get(
            f"{SUPABASE_URL}/rest/v1/lawyers",
            params={"select": "id,name"},
            headers={**headers, "Prefer": ""},
        )
        resp.raise_for_status()
        lawyer_map = {l["name"]: l["id"] for l in resp.json()}
    print(f"  Supabase 律師：{len(lawyer_map)} 位")

    # 組裝 upsert 資料
    rows = []
    skipped = set()
    for _, row in grouped.iterrows():
        lawyer_id = lawyer_map.get(row["諮詢律師"])
        if not lawyer_id:
            skipped.add(row["諮詢律師"])
            continue
        rows.append({
            "lawyer_id": lawyer_id,
            "month": row["month"],
            "consult_count": int(row["consult_count"]),
            "signed_count": int(row["signed_count"]),
            "sign_rate": float(row["sign_rate"]),
            "revenue": int(row.get("revenue", 0)),
            "collected": int(row.get("collected", 0)),
        })

    if skipped:
        print(f"  找不到律師：{', '.join(sorted(skipped))}")

    if not rows:
        print("  沒有資料需要更新")
        return 0

    # Upsert (分批 50 筆)
    with httpx.Client(timeout=30) as client:
        upsert_headers = {
            **headers,
            "Prefer": "resolution=merge-duplicates,return=minimal",
        }
        batch_size = 50
        success = 0
        for i in range(0, len(rows), batch_size):
            batch = rows[i:i + batch_size]
            resp = client.post(
                f"{SUPABASE_URL}/rest/v1/monthly_stats",
                json=batch,
                headers=upsert_headers,
            )
            if resp.status_code in (200, 201):
                success += len(batch)
            else:
                print(f"  upsert 失敗：{resp.status_code} {resp.text[:200]}")

    print(f"  Supabase 已更新：{success}/{len(rows)} 筆月統計")

    # ── 同步個別案件到 consultation_cases ──
    case_rows = []
    for _, row in df.iterrows():
        lawyer_id = lawyer_map.get(row["諮詢律師"])
        if not lawyer_id:
            continue
        case_number = str(row.get("案件編號", "")).strip() if "案件編號" in df.columns else ""
        if not case_number or case_number == "nan":
            # 自動產生唯一編號：CRM_律師ID_日期_序號
            case_date_str = row["諮詢日期"].strftime("%Y-%m-%d") if hasattr(row["諮詢日期"], "strftime") else str(row["諮詢日期"])[:10]
            case_number = f"CRM_{lawyer_id[:8]}_{case_date_str}_{len(case_rows)}"
        sign_status = str(row.get("簽約狀態", "")).strip()
        is_signed = sign_status != "" and sign_status != "nan" and "未" not in sign_status
        client_name = str(row.get("當事人", "")).strip() if "當事人" in df.columns else ""
        if client_name == "nan":
            client_name = ""
        case_type_col = next((c for c in df.columns if "服務項目" in c), None)
        case_type = str(row.get(case_type_col, "")).strip() if case_type_col else ""
        if case_type == "nan":
            case_type = ""
        case_revenue = int(float(row.get("revenue", 0) or 0)) if "revenue" in df.columns else 0
        case_collected = int(float(row.get("collected", 0) or 0)) if "collected" in df.columns else 0
        case_rows.append({
            "lawyer_id": lawyer_id,
            "case_date": row["諮詢日期"].strftime("%Y-%m-%d") if hasattr(row["諮詢日期"], "strftime") else str(row["諮詢日期"])[:10],
            "case_type": case_type,
            "case_number": case_number,
            "client_name": client_name,
            "is_signed": is_signed,
            "revenue": case_revenue,
            "collected": case_collected,
        })

    # Debug: 統計金額
    cases_with_amount = [r for r in case_rows if r.get("collected", 0) > 0 or r.get("revenue", 0) > 0]
    print(f"  consultation_cases 準備寫入：{len(case_rows)} 筆，其中有金額 {len(cases_with_amount)} 筆")
    if cases_with_amount:
        print(f"    金額範例：{cases_with_amount[0]}")
    elif case_rows:
        print(f"    ⚠ 全部金額為 0，範例：{case_rows[0]}")

    if case_rows:
        with httpx.Client(timeout=30) as client:
            case_headers = {
                **headers,
                "Prefer": "resolution=merge-duplicates,return=minimal",
            }
            case_success = 0
            for i in range(0, len(case_rows), 50):
                batch = case_rows[i:i + 50]
                resp = client.post(
                    f"{SUPABASE_URL}/rest/v1/consultation_cases",
                    json=batch,
                    headers=case_headers,
                )
                if resp.status_code in (200, 201):
                    case_success += len(batch)
                else:
                    print(f"  consultation_cases upsert 失敗：{resp.status_code} {resp.text[:200]}")
        print(f"  consultation_cases 已更新：{case_success}/{len(case_rows)} 筆")

    return success


# ═══════════════════════════════════════════════════════════
#  Step 3: 寫入同步狀態
# ═══════════════════════════════════════════════════════════

def write_sync_status(status, message, scraped_months="", rows_scraped=0, rows_updated=0, started_at=None):
    """將同步結果寫入 sync_status 表（用 service_role key 繞過 RLS）。"""
    import httpx

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    payload = {
        "id": "daily_update",
        "status": status,
        "message": message,
        "scraped_months": scraped_months,
        "rows_scraped": rows_scraped,
        "rows_updated": rows_updated,
        "started_at": started_at,
        "finished_at": datetime.utcnow().isoformat() + "Z",
        "updated_at": datetime.utcnow().isoformat() + "Z",
    }
    try:
        with httpx.Client(timeout=10) as client:
            resp = client.post(
                f"{SUPABASE_URL}/rest/v1/sync_status",
                json=payload,
                headers=headers,
            )
            if resp.status_code not in (200, 201):
                print(f"  (同步狀態寫入失敗: {resp.status_code})")
    except Exception as e:
        print(f"  (同步狀態寫入失敗: {e})")


# ═══════════════════════════════════════════════════════════
#  主程式
# ═══════════════════════════════════════════════════════════

def resolve_target_months(args):
    """根據命令列參數決定要更新哪些月份。"""
    now = datetime.now()

    if args.all:
        # 2020 至今
        months = set()
        for y in range(2020, now.year + 1):
            end_m = now.month if y == now.year else 12
            for m in range(1, end_m + 1):
                months.add((y, m))
        return months

    if args.month:
        # 指定月份，例如 2026-03
        y, m = map(int, args.month.split("-"))
        return {(y, m)}

    if args.months:
        # 最近 N 個月
        months = set()
        y, m = now.year, now.month
        for _ in range(args.months):
            months.add((y, m))
            m -= 1
            if m == 0:
                m = 12
                y -= 1
        return months

    # 預設：本月
    return {(now.year, now.month)}


def main():
    parser = argparse.ArgumentParser(description="每日一鍵更新：CRM 爬取 + Supabase 更新")
    parser.add_argument("--month", help="指定月份 (格式: 2026-03)")
    parser.add_argument("--months", type=int, help="最近 N 個月")
    parser.add_argument("--all", action="store_true", help="全部重抓 (2020 至今)")
    parser.add_argument("--skip-scrape", action="store_true", help="跳過爬蟲，只更新 Supabase")
    parser.add_argument("--skip-supabase", action="store_true", help="跳過 Supabase，只爬 CRM")
    args = parser.parse_args()

    target_months = resolve_target_months(args)
    sorted_months = sorted(target_months)
    first = sorted_months[0]
    last = sorted_months[-1]

    print()
    print("╔══════════════════════════════════════════════╗")
    print("║   喆律儀表板 - 每日資料更新                  ║")
    print("╚══════════════════════════════════════════════╝")
    print(f"  更新範圍：{first[0]}/{first[1]:02d} ~ {last[0]}/{last[1]:02d}（{len(target_months)} 個月）")

    # 取得 CRM 帳密
    email = CRM_USERNAME
    password = CRM_PASSWORD
    if not args.skip_scrape:
        if not email:
            email = input("  CRM 帳號：").strip()
        if not password:
            import getpass
            password = getpass.getpass("  CRM 密碼：")

    month_label = f"{first[0]}/{first[1]:02d} ~ {last[0]}/{last[1]:02d}"
    started_at = datetime.utcnow().isoformat() + "Z"
    rows_scraped = 0
    rows_updated = 0

    # 寫入「執行中」狀態
    write_sync_status("running", f"正在更新 {month_label}...", month_label, started_at=started_at)

    try:
        # Step 1: 爬 CRM
        if not args.skip_scrape:
            rows_scraped = scrape_crm(target_months, email, password)
        else:
            print("\n  (跳過 CRM 爬蟲)")

        # Step 2: 更新 Supabase
        if not args.skip_supabase:
            result = update_supabase(target_months)
            rows_updated = result or 0
        else:
            print("\n  (跳過 Supabase 更新)")

    except Exception as e:
        print(f"\n  ❌ 錯誤：{e}")
        import traceback
        traceback.print_exc()
        write_sync_status("error", f"更新失敗：{str(e)[:200]}", month_label,
                          rows_scraped, rows_updated, started_at)
        sys.exit(1)

    # 寫入成功狀態
    write_sync_status("success", f"更新完成 {month_label}", month_label,
                      rows_scraped, rows_updated, started_at)

    print(f"\n{'='*50}")
    print("  全部完成！")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
