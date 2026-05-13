"""
analyze_judicial_cross_referral_v2.py
------------------------------------
v2：直接從司法官 Excel 收入 sheet 的 col 3（諮詢律師欄）判定跨轉案，
不再反查 consultation_cases（更可靠：Excel 是分潤的 source of truth）。

判定規則：
  col 3 = 司法官本人                  → self_consult
  col 3 = 其他司法官                  → cross_judicial
  col 3 是其他律師（北所/合署/主持）  → cross_other  ←【喆律轉的】
  col 3 空白                          → no_consult_filled (多半是舊客續委或漏填)
  多位諮詢律師（逗號分隔）            → 取**第一位**做判定

輸出：
  judicial_cross_referral_v2.csv
"""
from __future__ import annotations

import csv, io, os, re, sys
from collections import defaultdict
from pathlib import Path
from datetime import datetime
import openpyxl

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

INPUT_DIRS = [
    r"C:\Users\admin\Desktop\新增資料夾\drive-download-20260418T020634Z-3-001",
]
FILENAME_PATTERN = re.compile(r"(\d{3})年.*?(方心瑜|孫少輔|許致維|劉明潔)律師案件明細")
JUDICIAL = {"劉明潔", "方心瑜", "孫少輔", "許致維"}
# 合署內部諮詢人力（諮詢→司法官承辦不算跨轉，算 cohort 內部消化）
COHORT_CONSULTANTS = {"曾秉浩", "劉誠夫"}
COHORT_FULL = JUDICIAL | COHORT_CONSULTANTS
OUTPUT_CSV = Path(r"C:\projects\lawyer-dashboard\judicial_cross_referral_v2.csv")


def parse_sheet_name(name: str, default_year: int) -> tuple[int | None, int | None]:
    """sheet name → (year, month) ROC. 只回有「收入」/「分潤」之外不管。"""
    digits = re.findall(r"\d+", name)
    for d in digits:
        if len(d) == 5:
            y, mo = int(d[:3]), int(d[3:])
            if 100 <= y <= 120 and 1 <= mo <= 12:
                return y, mo
    m = re.search(r"(\d{3})年\D*(\d{1,2})月", name)
    if m:
        return int(m.group(1)), int(m.group(2))
    m2 = re.search(r"(\d{1,2})月", name)
    if m2:
        mo = int(m2.group(1))
        if 1 <= mo <= 12:
            return default_year, mo
    return None, None


def first_consult_name(raw) -> str:
    """col 3 可能是 '劉誠夫, 孫少輔'，取第一位。"""
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    # 拆逗號 / 頓號
    parts = re.split(r"[,、，]", s)
    for p in parts:
        p = p.strip()
        if p:
            return p
    return ""


def is_lawyer_name(s: str) -> bool:
    """看起來像律師名字（2-4 個中文字、不是支付方式/狀態字串）"""
    if not s:
        return False
    if len(s) > 6 or len(s) < 2:
        return False
    bad = {"轉帳", "現金", "退款", "付款", "新客", "舊客", "是", "否", "未填"}
    if s in bad:
        return False
    return True


def find_files():
    out = []
    for d in INPUT_DIRS:
        p = Path(d)
        if not p.exists():
            continue
        for f in p.glob("*.xlsx"):
            m = FILENAME_PATTERN.search(f.name)
            if not m:
                continue
            out.append((f, int(m.group(1)), m.group(2)))
    return sorted(out, key=lambda x: (x[2], x[1]))


def parse_income_sheet(ws, lawyer: str, year: int, month: int):
    """讀收入 sheet → 每筆 case (client, consult_lawyer, amount, date, section, voided)。"""
    rows = []
    section = "承辦"
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # section header (col 0 = '自案'/'介紹'/'承辦'/'XX明細')
        first = str(row[0]).strip() if row[0] is not None else ""
        if first:
            for kw in ("自案", "介紹", "受僱", "其他"):
                if kw in first:
                    section = kw
                    break
            if "承辦" in first or "明細" in first:
                section = "承辦"

        # data row needs: client (col 1) 跟 amount 都有
        client = row[1] if len(row) > 1 else None
        amount = row[5] if len(row) > 5 else None
        date_val = row[6] if len(row) > 6 else None
        consult = row[3] if len(row) > 3 else None
        voided = row[12] if len(row) > 12 else None

        if client is None or not isinstance(amount, (int, float)):
            continue
        client_s = str(client).strip()
        if not client_s:
            continue
        if client_s in ("小計", "合計", "總計", "姓名", "當事人"):
            continue

        # date 正規化
        if isinstance(date_val, datetime):
            date_s = date_val.strftime("%Y-%m-%d")
        elif date_val is not None:
            date_s = str(date_val)[:10]
        else:
            date_s = ""

        consult_first = first_consult_name(consult)
        if not is_lawyer_name(consult_first):
            consult_first = ""

        rows.append({
            "judicial": lawyer,
            "year": year,
            "month": month,
            "section": section,
            "client": client_s,
            "consult_lawyer": consult_first,
            "consult_raw": str(consult).strip() if consult else "",
            "amount": float(amount),
            "date": date_s,
            "voided": str(voided).strip() if voided else "",
        })
    return rows


def classify(consult: str, judicial: str) -> str:
    if not consult:
        return "no_consult_filled"
    if consult == judicial:
        return "self_consult"
    if consult in JUDICIAL:
        return "cross_judicial"
    if consult in COHORT_CONSULTANTS:
        return "cohort_internal"   # 合署內部諮詢人力（曾秉浩、劉誠夫）
    return "cross_other"            # 真正的喆律端轉案（非合署）


def main():
    files = find_files()
    print(f"Found {len(files)} files")
    all_rows = []
    for f, year, lawyer in files:
        wb = openpyxl.load_workbook(f, data_only=True)
        for sn in wb.sheetnames:
            if "收入" not in sn or "副本" in sn:
                continue
            sy, sm = parse_sheet_name(sn, year)
            if sy is None:
                continue
            ws = wb[sn]
            rows = parse_income_sheet(ws, lawyer, sy, sm)
            all_rows.extend(rows)
        wb.close()
    print(f"Parsed {len(all_rows)} rows")

    # 過濾承辦 section + 未作廢
    承辦 = [r for r in all_rows if r["section"] == "承辦" and r["voided"] != "是"]
    print(f"承辦 + 未作廢: {len(承辦)}")

    for r in 承辦:
        r["category"] = classify(r["consult_lawyer"], r["judicial"])

    # 摘要
    by_cat_cnt = defaultdict(int)
    by_cat_amt = defaultdict(float)
    by_jud = defaultdict(lambda: defaultdict(int))
    by_src = defaultdict(lambda: defaultdict(float))  # consult_lawyer → {cnt, amt}
    for r in 承辦:
        c = r["category"]
        by_cat_cnt[c] += 1
        by_cat_amt[c] += r["amount"]
        by_jud[r["judicial"]][c] += 1
        if c in ("cross_other", "cross_judicial"):
            by_src[r["consult_lawyer"]]["cnt"] += 1
            by_src[r["consult_lawyer"]]["amt"] += r["amount"]

    print("\n=== 分類（v2.1：扣除合署內部諮詢人力曾秉浩/劉誠夫）===")
    for k in ["self_consult", "cohort_internal", "cross_other", "cross_judicial", "no_consult_filled"]:
        print(f"  {k:20s} {by_cat_cnt[k]:>5}  ${by_cat_amt[k]:>15,.0f}")

    print("\n=== by 司法官 ===")
    for j in sorted(JUDICIAL):
        d = by_jud[j]
        print(f"  {j}: self={d['self_consult']:>3}  cohort={d['cohort_internal']:>3}  "
              f"cross_other={d['cross_other']:>3}  no_fill={d['no_consult_filled']:>3}")

    print("\n=== 跨轉來源律師 Top 15（依件數） ===")
    top = sorted(by_src.items(), key=lambda kv: -kv[1]["cnt"])[:15]
    for name, d in top:
        print(f"  {name:10s} {int(d['cnt']):>4}  ${d['amt']:>12,.0f}")

    # 寫 CSV
    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_CSV, "w", encoding="utf-8-sig", newline="") as fp:
        w = csv.DictWriter(fp, fieldnames=[
            "category", "judicial", "year", "month", "date",
            "client", "consult_lawyer", "consult_raw",
            "amount", "section",
        ])
        w.writeheader()
        for r in 承辦:
            w.writerow({k: r.get(k) for k in w.fieldnames})
    print(f"\n✓ wrote {OUTPUT_CSV} ({len(承辦)} rows)")

    # 印 cross_other top 30 by amount
    print("\n=== cross_other 明細（前 30 筆，依金額） ===")
    co = [r for r in 承辦 if r["category"] == "cross_other"]
    co.sort(key=lambda r: -r["amount"])
    for r in co[:30]:
        print(f"  {r['judicial']:6s} ← {r['consult_lawyer']:10s}  {r['date']}  "
              f"{r['client']:14s}  ${r['amount']:>10,.0f}")


if __name__ == "__main__":
    main()
