#!/usr/bin/env python3
"""
律師年資對照 ETL — 產 public/forecast/lawyer-tenure.json
供五年推估頁的「合署轉換」lever 精算「在職 ≥N 年、非主管、尚未轉合署」的可轉律師。

資料源：
  - 到職日：人事薪資異動筆記.xlsx「1-入職」分頁（offer 欄日期最一致，民國/ISO 混格式）
  - 離職：「8-離職」分頁 + lawyers.departed_at
  - 現職律師名單 / 所別：LAWYERS_BY_OFFICE（與 build_case_cost_data.py 對齊，source of truth）
  - 已轉合署：PARTNER_SINCE（與 build_case_cost_data.py 對齊）→ 從可轉池排除
  - 年薪：finance_employees_monthly 最新民國年（115）×折年
  - 年 billing：consultation_cases 已簽約 booking by lawyer（2025），缺則 null（前端用 slider 預設補）

口徑警示：
  - 只保留「現職律師名單」中的人（Excel 含大量已離職/實習者，不可全進）
  - 主管(MANAGERS) 與 非諮詢(NON_CONSULTING) 標記排除
  - hire_year 取 offer 年；Excel 找不到者（多為 2019 前資深創所律師）標 hire_year=2018（年資已達標）
"""
import os, sys, io, json, glob, re
from collections import defaultdict
from datetime import datetime, date

import requests
from dotenv import load_dotenv
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
load_dotenv('scripts/.env')
SB_URL = (os.environ.get('SUPABASE_URL') or 'https://zpbkeyhxyykbvownrngf.supabase.co').rstrip('/')
KEY = os.environ['SUPABASE_SERVICE_KEY']
H = {'apikey': KEY, 'Authorization': f'Bearer {KEY}'}

# ── 現職律師名單（與 build_case_cost_data.py LAWYERS_BY_OFFICE 對齊）──
LAWYERS_BY_OFFICE = {
    '台北所': {'雷皓明','黃杰','孫少輔','許致維','劉明潔','方心瑜','張又仁','林桑羽','黃顯皓','柯雪莉',
        '陳寧馨','林昀','張嘉淳','黃世欣','李家泓','徐品軒','蘇萱','林宜嫻','吳柏慶','蕭予馨',
        '徐棠娜','劉誠夫','陳俊瑋','王怡婷','曾秉浩','李育哲','楊典翰','莊喬鈞','楊啓廷','張文祈',
        '劉庭懿','秦薇妮','黃庭汶','陳彥銘','陳昱璇','葉欣瑩','謝宗蓉','林敬修'},
    '桃園所': {'李杰峰','嚴心吟','張元毓','劉雅涵','李家徹','張佳榕','林品妘','王相為','王相爲'},
    '新竹所': {'陶光星','張家瑜','楊睿杰','葉芷羽'},
    '台中所': {'洪琬琪','李昭萱','許煜婕','陳璽仲','林佳穎','劉奕靖','李佳蓉','黃子菱'},
    '台南所': {'王湘閔','黃馨儀','黃書炫','姜奕成'},
    '高雄所': {'王郁萱','廖懿涵','陳映臻','蘇端雅'},
}
NAME_TO_OFFICE = {n: o for o, names in LAWYERS_BY_OFFICE.items() for n in names}
CURRENT_LAWYERS = set(NAME_TO_OFFICE)

# 已轉合署（build_case_cost_data.py PARTNER_SINCE）→ 從可轉池排除
ALREADY_PARTNER = {
    '孫少輔','許致維','劉明潔','方心瑜','陳璽仲','許煜婕','蕭予馨','徐棠娜','林昀','李昭萱',
    '柯雪莉','吳柏慶','黃顯皓','蘇萱','黃世欣','劉誠夫','陳俊瑋','曾秉浩',
}
# 主管 / 非諮詢（除主管外才轉合署）
#   雷皓明=主持律師；張飛宇=財務主管；張又仁、林桑羽=台北所主管（使用者指定）
#   各所所長（桃園李杰峰/新竹陶光星/台中洪琬琪/台南王湘閔/高雄王郁萱）= 各所最資深，使用者確認可換
MANAGERS = {
    '雷皓明', '黃杰', '張飛宇', '張又仁', '林桑羽',
    '李杰峰', '陶光星', '洪琬琪', '王湘閔', '王郁萱',
}

ROC_OFFSET = 1911


def parse_year(v):
    """從 offer/到職 欄取西元年。支援 datetime、民國 '0109/11/18'、ISO '2022-04-13'。"""
    if v is None: return None
    if isinstance(v, (datetime, date)):
        return v.year
    s = str(v).strip()
    if not s or s in ('X', 'Ｘ', 'x'): return None
    m = re.match(r'^0?(1\d{2})[/\-.]', s)        # 民國 109/110/111...
    if m:
        return int(m.group(1)) + ROC_OFFSET
    m = re.match(r'^(20\d{2})[-/]', s)            # ISO 2022-...
    if m:
        return int(m.group(1))
    m = re.match(r'^(1\d{2})$', s)                # 純民國年
    if m:
        return int(m.group(1)) + ROC_OFFSET
    return None


# ── 1. 解析 Excel 入職 ──
print('[1/4] 解析 Excel 入職/離職 ...')
xls = glob.glob(os.path.expanduser('~/Downloads/人事薪資異動筆記*.xlsx'))
if not xls:
    xls = glob.glob('人事薪資異動筆記*.xlsx')
if not xls:
    print('ERROR: 找不到 人事薪資異動筆記*.xlsx（請放 ~/Downloads/）', file=sys.stderr); sys.exit(1)
wb = openpyxl.load_workbook(xls[0], read_only=True, data_only=True)

hire_year = {}   # name -> 西元年
ws = wb['1-入職']
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0 or not row or len(row) < 6: continue
    name = (str(row[4]).strip() if row[4] else '')
    pos = (str(row[5]) if row[5] else '')
    if not name or '律師' not in pos: continue
    y = parse_year(row[0]) or parse_year(row[1])   # offer 優先, 到職 fallback
    if y and (name not in hire_year or y < hire_year[name]):
        hire_year[name] = y

depart_year = {}
ws = wb['8-離職']
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0 or not row or len(row) < 5: continue
    name = (str(row[2]).strip() if row[2] else '')
    if not name: continue
    y = parse_year(row[4])   # 離職日
    if y: depart_year[name] = y
print(f'    Excel 律師到職: {len(hire_year)} 筆；離職: {len(depart_year)} 筆')

# ── 2. lawyers.departed_at 補離職 ──
print('[2/4] lawyers.departed_at ...')
for r in requests.get(f'{SB_URL}/rest/v1/lawyers?select=name,departed_at', headers=H, timeout=30).json():
    da = r.get('departed_at')
    if da:
        try: depart_year[(r.get('name') or '').strip()] = int(str(da)[:4])
        except Exception: pass

# ── 3. 年薪（finance_employees_monthly 最新民國年 115）──
print('[3/4] 年薪 finance_employees_monthly ...')
sal_rows = requests.get(f'{SB_URL}/rest/v1/finance_employees_monthly?select=fiscal_year,month,name,salary_subtotal&fiscal_year=eq.115', headers=H, timeout=40).json()
sal_sum = defaultdict(float); sal_months = defaultdict(set)
for r in sal_rows:
    nm = (r.get('name') or '').strip()
    sal_sum[nm] += float(r.get('salary_subtotal') or 0); sal_months[nm].add(int(r['month']))
annual_salary = {nm: (sal_sum[nm] / len(sal_months[nm]) * 12) for nm in sal_sum if sal_months[nm]}

# ── 4. 年 billing（consultation_cases 已簽約 booking 2025, lawyer_id→name）──
print('[4/4] 年 billing consultation_cases(2025) ...')
lawyer_id_to_name = {}
off = 0
while True:
    chunk = requests.get(f'{SB_URL}/rest/v1/lawyers?select=id,name&limit=1000&offset={off}', headers=H, timeout=30).json()
    for r in chunk:
        if r.get('id'): lawyer_id_to_name[r['id']] = (r.get('name') or '').strip()
    if len(chunk) < 1000: break
    off += 1000
billing = defaultdict(float); off = 0
while True:
    chunk = requests.get(f'{SB_URL}/rest/v1/consultation_cases?select=case_date,is_signed,revenue,lawyer_id&limit=1000&offset={off}', headers=H, timeout=40).json()
    for r in chunk:
        if not r.get('is_signed') or not r.get('revenue'): continue
        cd = r.get('case_date')
        if not cd or str(cd)[:4] != '2025': continue
        nm = lawyer_id_to_name.get(r.get('lawyer_id'))
        if nm: billing[nm] += float(r['revenue'])
    if len(chunk) < 1000: break
    off += 1000

# ── 組裝 roster（只保留現職律師名單）──
THIS_YEAR = 2026
lawyers = []
n_hire_excel = n_hire_fallback = 0
for name in sorted(CURRENT_LAWYERS):
    hy = hire_year.get(name)
    if hy: n_hire_excel += 1
    else: hy = 2018; n_hire_fallback += 1   # 名單在但 Excel 無 → 多為 2019 前資深，年資已達標
    lawyers.append({
        'name': name,
        'office': NAME_TO_OFFICE.get(name),
        'hire_year': hy,
        'hire_source': 'excel' if hire_year.get(name) else 'fallback_pre2019',
        'is_lawyer': True,
        'is_manager': name in MANAGERS,
        'already_partner': name in ALREADY_PARTNER,
        'depart_year': depart_year.get(name),
        'annual_salary': round(annual_salary.get(name, 0)) or None,
        'annual_billing': round(billing.get(name, 0)) or None,
        'billing_source': 'consultation_cases_2025' if billing.get(name) else None,
    })

# 可轉池彙總：在職、非主管、未轉合署、tenure>=threshold 的「池總額」（不輸出個人，避免 PII）
def pool_aggregate(year, threshold):
    cnt = cnt_bill = cnt_sal = 0
    sum_bill = sum_sal = 0
    for L in lawyers:
        active = (not L['depart_year']) or L['depart_year'] >= year
        if not (active and L['is_lawyer'] and not L['is_manager'] and not L['already_partner']
                and (year - L['hire_year']) >= threshold):
            continue
        cnt += 1
        if L['annual_billing']:
            cnt_bill += 1; sum_bill += L['annual_billing']
        if L['annual_salary']:
            cnt_sal += 1; sum_sal += L['annual_salary']
    return {'count': cnt, 'count_billing': cnt_bill, 'sum_billing': round(sum_bill),
            'count_salary': cnt_sal, 'sum_salary': round(sum_sal)}

# threshold 1..6 × 推估年 base+1..base+6 的池彙總（公開版只含這張表，無個人列）
FC_YEARS = list(range(THIS_YEAR, THIS_YEAR + 6))
eligibility = {str(th): {str(y): pool_aggregate(y, th) for y in FC_YEARS} for th in range(1, 7)}
summary = {str(y): eligibility['3'][str(y)]['count'] for y in FC_YEARS}   # tenure>=3 達標數（顯示用）

meta = {
    'generated_at': datetime.utcnow().isoformat() + 'Z',
    'source_xlsx': os.path.basename(xls[0]),
    'roster_basis': 'LAWYERS_BY_OFFICE（現職律師名單，與 build_case_cost_data.py 對齊）',
    'note': 'hire_source=fallback_pre2019 者為名單在但 Excel(2020起) 無到職紀錄，視為 2019 前資深律師（年資已達標）；annual_billing 缺者前端用 slider 預設補',
    'managers_excluded_count': len(MANAGERS & CURRENT_LAWYERS),
    'already_partner_excluded_count': len(ALREADY_PARTNER & CURRENT_LAWYERS),
    'anonymized': True,
}

# 公開版：只輸出「可轉池彙總表」(threshold×年 的 count/sum_billing/sum_salary)，
# 不含任何個人列 — 避免把個別律師薪資/billing 推到 public GitHub Pages（financial PII）。
meta['public_scope'] = '僅池級彙總(無個人列)；個別資料留本地'
meta['total_current_lawyers'] = len(lawyers)
output = {'meta': meta, 'base_year': THIS_YEAR - 1, 'eligible_by_year': summary, 'eligibility': eligibility}

os.makedirs('public/forecast', exist_ok=True)
with open('public/forecast/lawyer-tenure.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, separators=(',', ':'))

# 本地預覽（含姓名+個人，供核對；不進 public、不 commit）
with open('scripts/_tenure_preview.json', 'w', encoding='utf-8') as f:
    json.dump({'meta': {**meta, 'managers': sorted(MANAGERS & CURRENT_LAWYERS),
               'already_partner': sorted(ALREADY_PARTNER & CURRENT_LAWYERS)},
               'eligibility': eligibility, 'lawyers': lawyers}, f, ensure_ascii=False, indent=2)

print('\n══════ 律師年資 roster ══════')
print(f'  現職律師 {len(lawyers)} 位（Excel 到職 {n_hire_excel} / fallback {n_hire_fallback}）')
print(f'  已轉合署排除 {len(ALREADY_PARTNER & CURRENT_LAWYERS)} 位、主管排除 {len(MANAGERS & CURRENT_LAWYERS)} 位')
print(f'  可轉池（在職非主管未轉、tenure>=3）每年達標:')
for y, n in summary.items():
    print(f'    {y}: {n} 位')
have_sal = sum(1 for L in lawyers if L['annual_salary'])
have_bill = sum(1 for L in lawyers if L['annual_billing'])
print(f'  有年薪資料 {have_sal} 位 / 有 billing 資料 {have_bill} 位')
print(f'\nsaved: public/forecast/lawyer-tenure.json')
