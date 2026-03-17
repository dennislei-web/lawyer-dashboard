"""
create_template.py
產生每月案件上傳用的 Excel 範本。

使用方式：
  python create_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

TEMPLATE_PATH = Path(__file__).parent / "templates" / "monthly_upload_template.xlsx"


def create_template():
    wb = Workbook()

    # ── Sheet 1: 案件清單 ──
    ws = wb.active
    ws.title = "案件清單"

    headers = [
        ("律師姓名", 14),
        ("諮詢日期", 14),
        ("案件類型", 18),
        ("成案與否", 12),
        ("會議記錄檔名", 45),
        ("逐字稿檔名", 45),
        ("備註", 20),
    ]

    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)

    for col_idx, (header_text, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        col_letter = chr(64 + col_idx)
        ws.column_dimensions[col_letter].width = width

    # 範例資料
    sample_rows = [
        [
            "雷皓明",
            "2026-04-01",
            "離婚",
            "成案",
            "雷皓明_成案_20260401_離婚(會議記錄).docx",
            "雷皓明_成案_20260401_離婚(逐字稿).docx",
            "",
        ],
        [
            "張家萍",
            "2026-04-03",
            "詐欺",
            "未成案",
            "張家萍_未成案_20260403_詐欺(會議記錄).docx",
            "",
            "當事人取消後續",
        ],
        [
            "雷皓明",
            "2026-04-10",
            "侵害配偶權",
            "成案",
            "",
            "",
            "檔案依命名規則自動偵測",
        ],
    ]

    sample_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for row_idx, row_data in enumerate(sample_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = sample_fill

    # 資料驗證：成案與否下拉選單（D2:D1000）
    dv = DataValidation(type="list", formula1='"成案,未成案"', allow_blank=False)
    dv.error = "請選擇「成案」或「未成案」"
    dv.errorTitle = "輸入錯誤"
    ws.add_data_validation(dv)
    dv.add("D2:D1000")

    # ── Sheet 2: 說明 ──
    ws2 = wb.create_sheet(title="說明")
    ws2.column_dimensions["A"].width = 80

    ws2.cell(row=1, column=1, value="每月諮詢案件上傳說明").font = Font(bold=True, size=16)

    instructions = [
        "",
        "步驟 1：在「案件清單」工作表中填寫每筆諮詢案件資料",
        "  - 律師姓名、諮詢日期、案件類型、成案與否為必填欄位",
        "  - 諮詢日期格式：2026-04-01（年-月-日）",
        "  - 成案與否請使用下拉選單選擇",
        "",
        "步驟 2：將會議記錄與逐字稿檔案放在同一個資料夾中",
        "",
        "步驟 3：檔案命名規則",
        "  格式：律師名_成案或未成案_日期_案件類型(會議記錄).docx",
        "  範例：雷皓明_成案_20260401_離婚(會議記錄).docx",
        "  範例：雷皓明_成案_20260401_離婚(逐字稿).docx",
        "",
        "步驟 4：執行匯入指令",
        "  python monthly_import.py --xlsx 案件清單.xlsx --dir 檔案資料夾/",
        "",
        "步驟 5：如果檔案嚴格遵循命名規則，Excel 中的 E、F 欄（檔名）可留空",
        "  腳本會自動根據律師名+日期+案件類型比對資料夾中的檔案",
        "",
        "備註：",
        "  - 支援 .docx 與 .txt 格式（.pdf 會跳過）",
        "  - 可先用 --dry-run 測試：python monthly_import.py --xlsx 案件.xlsx --dir 檔案/ --dry-run",
        "  - 也可以只給資料夾，不提供 Excel：python monthly_import.py --dir 檔案資料夾/",
    ]

    for i, line in enumerate(instructions, start=2):
        ws2.cell(row=i, column=1, value=line)

    # 儲存
    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(TEMPLATE_PATH))
    print(f"範本已產生: {TEMPLATE_PATH}")


if __name__ == "__main__":
    create_template()
