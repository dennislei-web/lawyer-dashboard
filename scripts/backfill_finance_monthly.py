"""
Backfill finance_employees_monthly from salary Excel files (111-115 民國年).

執行：
  python scripts/backfill_finance_monthly.py              # dry-run (印報告，不寫 DB)
  python scripts/backfill_finance_monthly.py --commit     # 真的寫入 DB
  python scripts/backfill_finance_monthly.py --year 114   # 只跑單一年度

寫入前會先 delete (fiscal_year, month) 範圍的舊資料，避免重覆。
"""
import argparse
import os
import re
import sys
from pathlib import Path

import httpx
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv(Path(__file__).parent / ".env")

SUPABASE_URL = os.environ["SUPABASE_URL"].rstrip("/")
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_KEY"]


def sb_delete(table, query):
    """DELETE via PostgREST. query is a filter like 'fiscal_year=in.(111,112)'."""
    url = f"{SUPABASE_URL}/rest/v1/{table}?{query}"
    r = httpx.delete(url, headers={
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Prefer": "return=minimal",
    }, timeout=60.0)
    r.raise_for_status()


def sb_insert(table, rows):
    """INSERT batch via PostgREST."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    r = httpx.post(url, headers={
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }, json=rows, timeout=120.0)
    if r.status_code >= 400:
        raise RuntimeError(f"insert {table} failed: {r.status_code} {r.text[:500]}")

FILES = {
    111: r"C:\Users\admin\Downloads\喆律111年度薪資 (1).xlsx",
    112: r"C:\Users\admin\Downloads\喆律112年度薪資 (1).xlsx",
    113: r"C:\Users\admin\Downloads\喆律113年度薪資 (1).xlsx",
    114: r"C:\Users\admin\Downloads\喆律114年薪資 (5).xlsx",
    115: r"C:\Users\admin\Downloads\喆律115年薪資 (4).xlsx",
}

# ─── 111 年 A 欄 group label 白名單（forward-fill 用） ─────────────
DEPT_LABEL_111 = {
    "北所律師", "北所法務", "北所行政", "工讀", "品牌部",
    "中所", "桃所", "雄所", "台南所", "新竹所", "客服", "法律010",
}

# ─── 部門 → 接案所 映射 ────────────────────────────────────────────
DEPT_TO_OFFICE = {
    # 111 年的粗分類
    "北所律師": "台北所", "北所法務": "台北所", "北所行政": "台北所", "工讀": "台北所",
    # 112+ 年的細分
    "北所(接案、行政、工讀)": "台北所",
    "北所010": "台北所", "北所吉他": "台北所", "北所四部": "台北所", "北所金貝殼": "台北所",
    "中所": "台中所",
    "南所": "台南所", "台南所": "台南所",
    "桃所": "桃園所",
    "竹所": "新竹所", "新竹所": "新竹所",
    "雄所": "高雄所",
    # 非接案所（office = NULL）
    "公司": None, "客服": None, "其他": None, "法顧": None, "法律010": None, "品牌部": None,
    "司法官合署": None,
}


def map_office(dept: str) -> str | None:
    if not dept:
        return None
    return DEPT_TO_OFFICE.get(dept.strip())


def is_valid_name(n) -> bool:
    if not n:
        return False
    s = str(n).strip()
    if not s:
        return False
    if re.fullmatch(r"\d[\d\s,.]*", s):
        return False
    if not re.search(r"[一-鿿a-zA-Z]", s):
        return False
    if any(k in s for k in ("小計", "合計", "外包", "姓名")):
        return False
    return True


def detect_columns_modern(rows):
    """For 112+ format: header has 部門, 姓名, 本薪 columns."""
    for r in range(min(3, len(rows))):
        row = rows[r]
        idx = {"name": -1, "dept": -1, "base": -1, "subtotal": -1,
               "total": -1, "labor": -1, "health": -1, "pension": -1, "emp_total": -1,
               "bonus": -1, "deduct": -1}
        for c, val in enumerate(row):
            h = (str(val) if val is not None else "").strip()
            if h == "姓名":
                idx["name"] = c
            elif h == "部門":
                idx["dept"] = c
            elif "本薪" in h:
                idx["base"] = c
            elif "薪資小計" in h:
                idx["subtotal"] = c
            elif h.startswith("合計") and "支出" not in h:
                idx["total"] = c
        if idx["name"] >= 0 and idx["base"] >= 0:
            return idx, r
    return None, -1


def detect_columns_111(rows):
    """111 年格式：no 部門 column header; B=姓名, D=本薪."""
    for r in range(min(3, len(rows))):
        row = rows[r]
        idx = {"name": -1, "dept": -1, "base": -1, "subtotal": -1, "total": -1}
        for c, val in enumerate(row):
            h = (str(val) if val is not None else "").strip()
            if h == "姓名":
                idx["name"] = c
            elif "本薪" in h:
                idx["base"] = c
            elif "薪資小計" in h:
                idx["subtotal"] = c
            elif h.startswith("合計") and "支出" not in h:
                idx["total"] = c
        if idx["name"] >= 0 and idx["base"] >= 0:
            return idx, r
    return None, -1


def dedup_by_name(rows):
    """同人同月份有多筆 row（主薪+補發）→ merge: 保留 base 最大的那筆，補空 dept。"""
    by_name = {}
    for r in rows:
        ex = by_name.get(r["name"])
        if ex is None:
            by_name[r["name"]] = r
        else:
            if r["base_salary"] > ex["base_salary"]:
                # 新 row 主薪較大，但保留舊 row 的 dept 如果新的為空
                if not r["dept_raw"] and ex["dept_raw"]:
                    r["dept_raw"] = ex["dept_raw"]
                by_name[r["name"]] = r
            else:
                if not ex["dept_raw"] and r["dept_raw"]:
                    ex["dept_raw"] = r["dept_raw"]
    return list(by_name.values())


def parse_modern_sheet(ws, year):
    """Parse a 112-115 monthly sheet. Return list of {name, dept, base, subtotal, total}."""
    rows = list(ws.iter_rows(values_only=True))
    ci, header_row = detect_columns_modern(rows)
    if ci is None:
        return []
    out = []
    for i in range(header_row + 1, len(rows)):
        row = rows[i]
        if ci["name"] >= len(row):
            continue
        name = row[ci["name"]]
        name = str(name).strip() if name is not None else ""
        try:
            base_raw = row[ci["base"]] if ci["base"] < len(row) else 0
            base = float(base_raw) if base_raw not in (None, "") else 0
        except (ValueError, TypeError):
            base = 0
        if not is_valid_name(name) or base < 10000:
            continue
        dept = ""
        if ci["dept"] >= 0 and ci["dept"] < len(row):
            d_raw = row[ci["dept"]]
            dept = str(d_raw).strip() if d_raw is not None else ""
        try:
            sub_raw = row[ci["subtotal"]] if ci["subtotal"] >= 0 and ci["subtotal"] < len(row) else None
            subtotal = float(sub_raw) if sub_raw not in (None, "") else base
        except (ValueError, TypeError):
            subtotal = base
        out.append({
            "name": name,
            "dept_raw": dept,
            "base_salary": int(round(base)),
            "salary_subtotal": int(round(subtotal)),
        })
    return dedup_by_name(out)


def parse_111_sheet(ws):
    """Parse a 111 monthly sheet. A column = dept group label (forward-fill via whitelist).

    A 欄 非白名單值 (e.g. '眷屬加保', '推薦人獎金', '和解金', 或人名) 視為調整 row：
    跳過該 row 不算員工，但保持 current_dept 不變（後續 A=None 的員工 row 沿用）。
    """
    rows = list(ws.iter_rows(values_only=True))
    ci, header_row = detect_columns_111(rows)
    if ci is None:
        return []
    out = []
    current_dept = None
    for i in range(header_row + 1, len(rows)):
        row = rows[i]
        a = row[0] if len(row) > 0 else None
        if a is not None and str(a).strip():
            label = str(a).strip()
            if label in DEPT_LABEL_111:
                current_dept = label
            else:
                # 非白名單值（眷屬加保/推薦人獎金/個人姓名等調整 row）：跳過但保持 current_dept
                continue
        # Parse employee row
        if current_dept is None:
            continue
        if ci["name"] >= len(row):
            continue
        name = row[ci["name"]]
        name = str(name).strip() if name is not None else ""
        try:
            base_raw = row[ci["base"]] if ci["base"] < len(row) else 0
            base = float(base_raw) if base_raw not in (None, "") else 0
        except (ValueError, TypeError):
            base = 0
        if not is_valid_name(name) or base < 10000:
            continue
        try:
            sub_raw = row[ci["subtotal"]] if ci["subtotal"] >= 0 and ci["subtotal"] < len(row) else None
            subtotal = float(sub_raw) if sub_raw not in (None, "") else base
        except (ValueError, TypeError):
            subtotal = base
        out.append({
            "name": name,
            "dept_raw": current_dept,
            "base_salary": int(round(base)),
            "salary_subtotal": int(round(subtotal)),
        })
    return dedup_by_name(out)


def collect_year(year, path):
    """Return {month: [employee dict]} for a single Excel file."""
    if not Path(path).exists():
        print(f"  ❌ 檔案不存在: {path}")
        return {}
    wb = load_workbook(path, data_only=True, read_only=True)
    is_111 = (year == 111)

    by_month = {}
    if is_111:
        # 111 年: sheet name '111年1月' or 'X月'
        for sn in wb.sheetnames:
            m = re.fullmatch(r"(?:\d{3}年)?(\d{1,2})月", sn)
            if not m:
                continue
            mn = int(m.group(1))
            if not (1 <= mn <= 12):
                continue
            ws = wb[sn]
            employees = parse_111_sheet(ws)
            by_month[mn] = employees
    else:
        # 112+: sheet name 'YYMM' digits
        for sn in wb.sheetnames:
            m = re.fullmatch(r"(\d{3,4})(\d{2})", sn)
            if not m:
                continue
            yr_part = int(m.group(1))
            mn = int(m.group(2))
            if not (1 <= mn <= 12):
                continue
            # year_part 不符 → 視為會計遺漏（e.g. 114 年 Excel 留著 113-12 的 sheet 沒換）
            if yr_part != year:
                print(f"  ⚠️  sheet '{sn}' year_part={yr_part} 跟檔案年度 {year} 不符 → 跳過（會計可能還沒匯入此月）")
                continue
            ws = wb[sn]
            by_month[mn] = parse_modern_sheet(ws, year)
    return by_month


def backfill_missing_dept(by_month):
    """For each (name) with missing dept in early months, fill from later months.
    Handles 112 年 1-2 月 header 沒有「部門」column 的情況。
    """
    # name → first non-empty dept across all months
    name_dept = {}
    for mn in sorted(by_month.keys()):
        for e in by_month[mn]:
            if e["dept_raw"] and e["name"] not in name_dept:
                name_dept[e["name"]] = e["dept_raw"]
    # Fill in missing dept
    backfilled = 0
    for mn, employees in by_month.items():
        for e in employees:
            if not e["dept_raw"] and e["name"] in name_dept:
                e["dept_raw"] = name_dept[e["name"]]
                backfilled += 1
    return backfilled


def build_rows(year, by_month, source_file):
    """Convert parsed data → list of finance_employees_monthly rows."""
    rows = []
    for mn, employees in by_month.items():
        for e in employees:
            office = map_office(e["dept_raw"])
            rows.append({
                "fiscal_year": year,
                "month": mn,
                "name": e["name"],
                "department": e["dept_raw"] or None,
                "base_salary": e["base_salary"],
                "bonus": 0,
                "deduct": 0,
                "salary_subtotal": e["salary_subtotal"],
                "source_file": source_file,
                "office": office,
            })
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--commit", action="store_true", help="真的寫入 DB（預設只 dry-run）")
    ap.add_argument("--year", type=int, default=None, help="只跑指定年度")
    args = ap.parse_args()

    targets = FILES if args.year is None else {args.year: FILES.get(args.year)}
    if args.year and not targets.get(args.year):
        print(f"❌ 找不到 {args.year} 年的檔案")
        return 1

    grand_total = 0
    for year, path in targets.items():
        print(f"\n{'='*70}")
        print(f"📅 {year} 年: {Path(path).name}")
        print(f"{'='*70}")

        by_month = collect_year(year, path)
        if not by_month:
            print(f"  ⚠️  解析不到任何月份")
            continue

        backfilled = backfill_missing_dept(by_month)
        if backfilled:
            print(f"  🔧 跨月 dept backfill: {backfilled} 筆（112 年 1-2 月或單月缺失）")

        rows = build_rows(year, by_month, Path(path).name)

        # Summary
        months_sorted = sorted(by_month.keys())
        print(f"  涵蓋月份: {months_sorted}")
        print(f"  逐月人數: " + ", ".join(f"{m}月={len(by_month[m])}" for m in months_sorted))

        # Office mapping summary
        office_count = {}
        no_office = 0
        for r in rows:
            if r["office"]:
                office_count[r["office"]] = office_count.get(r["office"], 0) + 1
            else:
                no_office += 1
        print(f"  office 對應: {office_count}  (無接案所: {no_office})")

        # Department coverage
        no_dept = sum(1 for r in rows if not r["department"])
        if no_dept:
            print(f"  ⚠️  無部門資料 row 數: {no_dept}")

        # Sample unmapped depts (raw → office NULL but dept exists)
        unmapped = sorted({r["department"] for r in rows if r["department"] and r["office"] is None})
        if unmapped:
            print(f"  無對應接案所的部門 (預期會是支援部門): {unmapped}")

        print(f"  ➡️  共 {len(rows)} 筆 row")
        grand_total += len(rows)

        if args.commit:
            sb_delete("finance_employees_monthly", f"fiscal_year=eq.{year}")
            BATCH = 500
            for i in range(0, len(rows), BATCH):
                chunk = rows[i:i+BATCH]
                sb_insert("finance_employees_monthly", chunk)
            print(f"  ✅ 已寫入 DB: 刪除 {year} 年舊資料 + 寫入 {len(rows)} 筆")

    print(f"\n{'='*70}")
    print(f"總計: {grand_total} 筆 row")
    print(f"模式: {'COMMIT (已寫入 DB)' if args.commit else 'DRY-RUN (未寫入)'}")
    print(f"{'='*70}")


if __name__ == "__main__":
    sys.exit(main() or 0)
