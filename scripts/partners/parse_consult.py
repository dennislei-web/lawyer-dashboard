"""
諮詢律師（黃顯皓型）案件明細 Excel → 乾淨 CSV
-----------------------------------------
輸入：drive_input 底下 "NNN年-NNN年黃顯皓律師案件明細.xlsx" 等
輸出：
  - consult_profit_share.csv   每月 profit entries（含 tier / ratio / 當事人）
  - consult_cases.csv          每月案件逐筆
  - consult_monthly_totals.csv 每月 喆律分潤 vs 律師分潤 總計
  - _consult_parse_issues.txt  解析異常

顯皓 Excel sheet 結構（每月一張，sheet 名 = 'YYMM' 例 11502）：
  Top: 案件明細（header row col 0 = '顯皓律師'，cols: 當事人/接案人員/金額/...）
  Bottom: 月度結算
    Row N: header「委任金額 | 諮詢場次 | 委任/場次(萬倍) | 獎金率 | 獎金 | _ | 顯皓承辦案件 | 委任費用 | 承辦分潤60% | 喆律分潤40%」
    Rows N+1..: 顯皓承辦案件（cols 7-10）+ 第一行同時包含當月 委任金額/場次（cols 1-3）
    Later: 「顯皓自案案件 | 委任費用 | 承辦分潤90% | 喆律分潤10%」 + 自案 rows

分潤規則（顯皓）：
  - 諮詢費（金額=2000）：100% 喆律
  - 月度委任費（諮詢成案）：律師獎金率：
      11410-11412：< 3 萬倍 → 3%；≥3 萬 → 5%；≥5 萬 → 8%
      ≥ 11501（保底取消）：< 3 萬倍 → 0%；≥3 → 5%；≥5 → 8%
  - 顯皓承辦案件：律師 60% / 喆律 40%
  - 顯皓自案案件：律師 90% / 喆律 10%
  - 月固定費 130,000（喆律給律師）— 寫進 monthly_totals 但不在 profit_share

Tier 命名：
  - '諮詢費'      ratio=0.0  律師 0% / 喆律 100%
  - '諮詢成案'    ratio=當月律師獎金率  (0.0 / 0.03 / 0.05 / 0.08)
  - '顯皓承辦'    ratio=0.6  律師 60% / 喆律 40%
  - '顯皓自案'    ratio=0.9  律師 90% / 喆律 10%
"""
import openpyxl
import argparse
import csv
import io
import os
import re
import sys
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ---------- 設定 ----------
_env_input = os.environ.get('PARTNERS_CONSULT_INPUT_DIRS', '').strip()
_env_output = os.environ.get('PARTNERS_OUTPUT_DIR', '').strip()
INPUT_DIRS = [d for d in _env_input.split(os.pathsep) if d] if _env_input else [
    r'C:\Users\admin\Desktop\新增資料夾\drive-download-20260419T022533Z-3-001',
]
CONSULT_LAWYERS = ['黃顯皓']
FILENAME_PATTERN = re.compile(r'(\d{3})年.*?(' + '|'.join(CONSULT_LAWYERS) + r')律師案件明細')
OUTPUT_DIR = Path(_env_output) if _env_output else Path('.')

MONTHLY_FIXED_COST = {'黃顯皓': 130000}


def find_input_files():
    """收集 Excel；同 (lawyer, roc_year) 出現多份時（如「(1).xlsx」副本），保留最新修改檔。"""
    candidates = []
    for d in INPUT_DIRS:
        p = Path(d)
        if not p.exists():
            continue
        for f in p.glob('*.xlsx'):
            m = FILENAME_PATTERN.search(f.name)
            if not m:
                continue
            # 排除「(N).xlsx」/ 「Converted」/「副本」 等明顯副本（保留乾淨主檔）
            if re.search(r'\((\d+)\)\.xlsx$', f.name) or 'Converted -' in f.name or '副本' in f.name:
                # 仍當候選，下面 dedupe 時若沒主檔才用
                pass
            roc_year = int(m.group(1))
            lawyer = m.group(2)
            candidates.append((f, roc_year, lawyer))
    # dedupe by (lawyer, roc_year)：先取「不含 (N) 副本標記」的；都有副本標記則取最新 mtime
    by_key = {}
    for f, y, l in candidates:
        key = (l, y)
        is_copy = bool(re.search(r'\((\d+)\)\.xlsx$', f.name)) or 'Converted -' in f.name or '副本' in f.name
        if key not in by_key:
            by_key[key] = (f, y, l, is_copy)
        else:
            cur_f, cur_y, cur_l, cur_is_copy = by_key[key]
            # 偏好非副本；都同等再比 mtime
            if cur_is_copy and not is_copy:
                by_key[key] = (f, y, l, is_copy)
            elif (not cur_is_copy and not is_copy) or (cur_is_copy and is_copy):
                if f.stat().st_mtime > cur_f.stat().st_mtime:
                    by_key[key] = (f, y, l, is_copy)
    results = [(f, y, l) for (f, y, l, _) in by_key.values()]
    return sorted(results, key=lambda x: (x[2], x[1]))


YYMM_RE = re.compile(r'^(\d{3})(\d{2})$')


def classify_sheet(sn):
    sn = sn.strip()
    if sn in ('規則', '綜合', '總匯'):
        return (None, None, 'summary')
    m = YYMM_RE.match(sn)
    if not m:
        return (None, None, f'unrecognised name: {sn}')
    y, mo = int(m.group(1)), int(m.group(2))
    if not (100 <= y <= 120 and 1 <= mo <= 12):
        return (None, None, f'year/month out of range: {sn}')
    return (y, mo, None)


def is_num(x):
    return isinstance(x, (int, float)) and not isinstance(x, bool)


def to_num(x, default=0.0):
    if x is None:
        return default
    if is_num(x):
        return float(x)
    try:
        return float(str(x).replace(',', ''))
    except Exception:
        return default


def calc_bonus_ratio(year, month, ratio_x_wan):
    """依顯皓規則算當月律師獎金率（諮詢成案部分）。"""
    if ratio_x_wan >= 5:
        return 0.08
    if ratio_x_wan >= 3:
        return 0.05
    # 保底 3% 機制：11410-11412
    if year == 114 and month in (10, 11, 12):
        return 0.03
    return 0.0


# ---------- 案件明細解析 ----------
CASE_COL_MAP = {
    '當事人': 'client',
    '接案人員': 'handlers',
    '金額': 'amount',
    '日期': 'date',
    '備註': 'note',
    '品牌': 'brand',
    '接案所': 'office',
    '部門': 'dept',
    '類型': 'case_type',
    '是否作廢': 'voided',
    '客戶來源': 'source',
}


def find_case_header(rows):
    """顯皓 sheet 案件 header 在 row 0 / col 0 = 'XX律師'。"""
    for i, row in enumerate(rows):
        if i > 5:
            break
        if not row:
            continue
        first = str(row[0]) if row[0] is not None else ''
        # 顯皓格式：col 0 = 'XX律師' 或 'XX律師案件' 都接受
        if not (first.endswith('律師') or first.endswith('律師案件') or '案件' in first):
            continue
        idx = {}
        for j, v in enumerate(row):
            if v is None:
                continue
            s = str(v).strip()
            if s in CASE_COL_MAP and CASE_COL_MAP[s] not in idx:
                idx[CASE_COL_MAP[s]] = j
        if 'client' in idx and 'amount' in idx:
            return (i, idx)
    return None


def parse_case_section(rows, lawyer, year, month):
    out = []
    found = find_case_header(rows)
    if not found:
        return out
    header_idx, idx = found
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        if not row or all(v is None or (isinstance(v, str) and not v.strip()) for v in row[:12]):
            # blank row → 結束 case 區
            if i - header_idx > 1:
                break
            continue
        client = row[idx['client']] if idx['client'] < len(row) else None
        if client is None:
            continue
        client_s = str(client).strip()
        if not client_s or client_s in ('小計', '合計', '總計', '當事人', '委任金額', '顯皓承辦案件'):
            break
        amt = row[idx['amount']] if 'amount' in idx and idx['amount'] < len(row) else None
        if isinstance(amt, str):
            try:
                amt = float(amt.replace(',', ''))
            except Exception:
                pass
        dt = row[idx['date']] if 'date' in idx and idx['date'] < len(row) else None
        date_str = dt.strftime('%Y-%m-%d') if isinstance(dt, datetime) else (str(dt) if dt is not None else None)

        def gv(f):
            j = idx.get(f)
            return row[j] if j is not None and j < len(row) else None

        out.append({
            'lawyer': lawyer, 'year': year, 'month': month,
            'section': '承辦',
            'client': client_s,
            'handlers': gv('handlers'),
            'amount': amt,
            'date': date_str,
            'note': gv('note'),
            'brand': gv('brand'),
            'office': gv('office'),
            'dept': gv('dept'),
            'case_type': gv('case_type'),
            'voided': gv('voided'),
            'source': gv('source'),
        })
    return out


# ---------- 結算解析 ----------
HANDLE_HEADER_RE = re.compile(r'承辦案件$')


def _find_string_col(row, target_substr):
    """在 row 裡找含特定字串的 col index，找不到回 None。"""
    if not row:
        return None
    for j, v in enumerate(row):
        if v is None:
            continue
        if target_substr in str(v):
            return j
    return None


def _find_exact_col(row, exact):
    if not row:
        return None
    for j, v in enumerate(row):
        if v is None:
            continue
        if str(v).strip() == exact:
            return j
    return None


def _parse_handle_block(rows, header_idx, client_col, kind, lawyer, year, month, ratio):
    """從 顯皓承辦/自案 header 行的下方讀 case data，直到空白/小計/應收/下一 header。
    kind = 'main' or 'self_case'; ratio 是律師分成比例。
    """
    entries = []
    amount_col = client_col + 1
    lawyer_col = client_col + 2
    zhelu_col = client_col + 3
    n_blank_in_a_row = 0
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        if not row:
            n_blank_in_a_row += 1
            if n_blank_in_a_row >= 2:
                break
            continue
        if client_col >= len(row):
            n_blank_in_a_row += 1
            if n_blank_in_a_row >= 2:
                break
            continue
        client_val = row[client_col]
        if client_val is None:
            n_blank_in_a_row += 1
            if n_blank_in_a_row >= 2:
                break
            continue
        n_blank_in_a_row = 0
        client_s = str(client_val).strip()
        if not client_s:
            continue
        # 終止：碰到下一個結算 header
        if client_s.endswith('承辦案件') or client_s.endswith('自案案件'):
            break
        # 終止：應收/合計/小計
        if (client_s.startswith('應收') or client_s in ('小計', '合計')):
            break
        # 跳過：標籤列再次出現「委任費用」之類
        if client_s in ('委任費用', '當事人'):
            continue
        amt_val = row[amount_col] if amount_col < len(row) else None
        if not is_num(amt_val):
            continue
        amt = to_num(amt_val)
        lawyer_amt = to_num(row[lawyer_col]) if lawyer_col < len(row) and is_num(row[lawyer_col]) else amt * ratio
        zhelu_amt = to_num(row[zhelu_col]) if zhelu_col < len(row) and is_num(row[zhelu_col]) else amt * (1 - ratio)
        tier = '顯皓承辦' if kind == 'main' else '顯皓自案'
        entries.append({
            'lawyer': lawyer, 'year': year, 'month': month,
            'side': 'lawyer_handled',
            'tier': tier,
            'client': client_s,
            'case_amount': amt,
            'ratio': ratio,
            'lawyer_amt': lawyer_amt,
            'zhelu_amt': zhelu_amt,
            'note': None,
        })
    return entries


def parse_settlement_section(rows, lawyer, year, month):
    """回傳 (profit_entries, monthly_consult_count, monthly_commission_total, bonus_ratio)。"""
    entries = []
    monthly_consult_count = 0
    monthly_commission_total = 0.0
    bonus_ratio = 0.0

    # Step 1: 找含「委任金額」字串的 row（commission header）
    commission_header_idx = None
    commission_cols = {}  # 'commission_total', 'consult_count', 'ratio_x_wan', 'bonus_ratio', 'bonus'
    for i, row in enumerate(rows):
        if row is None:
            continue
        if _find_exact_col(row, '委任金額') is None:
            continue
        commission_header_idx = i
        commission_cols['commission_total'] = _find_exact_col(row, '委任金額')
        commission_cols['consult_count'] = _find_exact_col(row, '諮詢場次')
        rxw_col = _find_string_col(row, '委任/場次')
        commission_cols['ratio_x_wan'] = rxw_col
        commission_cols['bonus_ratio'] = _find_exact_col(row, '獎金率')
        commission_cols['bonus'] = _find_exact_col(row, '獎金')
        break

    # Step 2: 從 header 下方找實際數值列（第一個在 commission_total 欄有數字的 row）
    if commission_header_idx is not None and commission_cols.get('commission_total') is not None:
        ct_col = commission_cols['commission_total']
        for i in range(commission_header_idx + 1, min(commission_header_idx + 6, len(rows))):
            row = rows[i]
            if not row or ct_col >= len(row):
                continue
            if is_num(row[ct_col]):
                monthly_commission_total = to_num(row[ct_col])
                cn_col = commission_cols.get('consult_count')
                if cn_col is not None and cn_col < len(row) and is_num(row[cn_col]):
                    monthly_consult_count = int(to_num(row[cn_col]))
                # 獎金率：優先取 Excel 寫好的，沒有就用我們算
                br_col = commission_cols.get('bonus_ratio')
                if br_col is not None and br_col < len(row) and is_num(row[br_col]):
                    bonus_ratio = to_num(row[br_col])
                    if bonus_ratio > 1:  # 若是 5 而非 0.05
                        bonus_ratio /= 100
                else:
                    rxw = (monthly_commission_total / (monthly_consult_count * 10000)
                           if monthly_consult_count > 0 else 0)
                    bonus_ratio = calc_bonus_ratio(year, month, rxw)
                break

    # Step 3: 找「顯皓承辦案件」 header → 解析下方
    main_header_idx = None
    main_client_col = None
    for i, row in enumerate(rows):
        if not row:
            continue
        for j, v in enumerate(row):
            if v is None:
                continue
            s = str(v).strip()
            if s.endswith('承辦案件') and '自案' not in s:
                main_header_idx = i
                main_client_col = j
                break
        if main_header_idx is not None:
            break

    if main_header_idx is not None:
        entries.extend(_parse_handle_block(rows, main_header_idx, main_client_col,
                                            'main', lawyer, year, month, ratio=0.6))

    # Step 4: 找「顯皓自案案件」header → 解析下方
    self_header_idx = None
    self_client_col = None
    for i, row in enumerate(rows):
        if not row:
            continue
        for j, v in enumerate(row):
            if v is None:
                continue
            s = str(v).strip()
            if s.endswith('自案案件'):
                self_header_idx = i
                self_client_col = j
                break
        if self_header_idx is not None:
            break

    if self_header_idx is not None:
        entries.extend(_parse_handle_block(rows, self_header_idx, self_client_col,
                                            'self_case', lawyer, year, month, ratio=0.9))

    return entries, monthly_consult_count, monthly_commission_total, bonus_ratio


# ---------- 主流程 ----------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--out-dir', default=str(OUTPUT_DIR), help='輸出資料夾')
    args = ap.parse_args()
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    files = find_input_files()
    if not files:
        # 沒檔案就產空 CSV，build_embedded.py 會 graceful skip
        for fn in ('consult_profit_share.csv', 'consult_cases.csv', 'consult_monthly_totals.csv'):
            (out_dir / fn).write_text('', encoding='utf-8-sig')
        print('  no consult input files; emitted empty CSVs')
        return 0

    cases_rows = []
    profit_rows = []
    monthly_totals = {}
    issues = []

    for fpath, roc_year, lawyer in files:
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
        except Exception as e:
            issues.append(f'CANNOT OPEN {fpath.name}: {e}')
            continue

        for sheet_name in wb.sheetnames:
            year, month, note = classify_sheet(sheet_name)
            if year is None:
                if note and 'summary' not in note:
                    issues.append(f'{fpath.name} :: {sheet_name}  -  skipped ({note})')
                continue
            if year != roc_year:
                pass

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            try:
                cases = parse_case_section(rows, lawyer, year, month)
                cases_rows.extend(cases)
            except Exception as e:
                issues.append(f'{fpath.name} :: {sheet_name}  case-parse failed: {e}')
                cases = []

            try:
                profit_entries, consult_count, commission_total, bonus_ratio = parse_settlement_section(
                    rows, lawyer, year, month)
            except Exception as e:
                issues.append(f'{fpath.name} :: {sheet_name}  settlement-parse failed: {e}')
                profit_entries = []
                consult_count = 0
                commission_total = 0.0
                bonus_ratio = 0.0

            # 1) 諮詢費 entries：從 case section 取 amount=2000 records
            consult_fee_entries = []
            for c in cases:
                if c.get('voided') == '是':
                    continue
                amt = c.get('amount')
                try:
                    amt_f = float(amt or 0)
                except (ValueError, TypeError):
                    continue
                if abs(amt_f) != 2000:
                    continue
                # 諮詢費：100% 喆律
                consult_fee_entries.append({
                    'lawyer': lawyer, 'year': year, 'month': month,
                    'side': 'zhelu_handled',
                    'tier': '諮詢費',
                    'client': c.get('client') or '',
                    'case_amount': amt_f,
                    'ratio': 0.0,
                    'lawyer_amt': 0.0,
                    'zhelu_amt': amt_f,
                    'note': None,
                })

            # 2) 月度委任費 entry（單筆匯總）
            commission_entry = None
            if commission_total > 0:
                lawyer_amt = commission_total * bonus_ratio
                zhelu_amt = commission_total - lawyer_amt
                commission_entry = {
                    'lawyer': lawyer, 'year': year, 'month': month,
                    'side': 'zhelu_handled',
                    'tier': '諮詢成案',
                    'client': '(月度委任匯總)',
                    'case_amount': commission_total,
                    'ratio': bonus_ratio,
                    'lawyer_amt': lawyer_amt,
                    'zhelu_amt': zhelu_amt,
                    'note': f'諮詢場次={consult_count}; 律師獎金率={bonus_ratio*100:.1f}%',
                }

            month_profits = list(consult_fee_entries)
            if commission_entry:
                month_profits.append(commission_entry)
            month_profits.extend(profit_entries)
            profit_rows.extend(month_profits)

            # 3) monthly totals
            zhelu_total = sum(e['zhelu_amt'] for e in month_profits)
            lawyer_total = sum(e['lawyer_amt'] for e in month_profits)
            fixed_cost = MONTHLY_FIXED_COST.get(lawyer, 0)
            # 月固定費：喆律端「淨收入」要扣，律師端「淨收入」要加
            net_zhelu = zhelu_total - fixed_cost
            net_lawyer = lawyer_total + fixed_cost

            monthly_totals[(lawyer, year, month)] = {
                'lawyer': lawyer, 'year': year, 'month': month,
                'zhelu_total': zhelu_total,
                'lawyer_total': lawyer_total,
                'fixed_cost': fixed_cost,
                'zhelu_net': net_zhelu,
                'lawyer_net': net_lawyer,
                'consult_count': consult_count,
                'commission_total': commission_total,
                'bonus_ratio': bonus_ratio,
            }

    # 寫 CSV
    profit_path = out_dir / 'consult_profit_share.csv'
    cases_path = out_dir / 'consult_cases.csv'
    totals_path = out_dir / 'consult_monthly_totals.csv'
    issues_path = out_dir / '_consult_parse_issues.txt'

    if profit_rows:
        with open(profit_path, 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.DictWriter(f, fieldnames=list(profit_rows[0].keys()))
            w.writeheader()
            w.writerows(profit_rows)
    else:
        profit_path.write_text('', encoding='utf-8-sig')

    if cases_rows:
        with open(cases_path, 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.DictWriter(f, fieldnames=list(cases_rows[0].keys()))
            w.writeheader()
            w.writerows(cases_rows)
    else:
        cases_path.write_text('', encoding='utf-8-sig')

    if monthly_totals:
        rows_out = list(monthly_totals.values())
        with open(totals_path, 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.DictWriter(f, fieldnames=list(rows_out[0].keys()))
            w.writeheader()
            w.writerows(rows_out)
    else:
        totals_path.write_text('', encoding='utf-8-sig')

    issues_path.write_text('\n'.join(issues) + '\n' if issues else '', encoding='utf-8')

    print(f'  wrote {len(profit_rows)} profit entries')
    print(f'  wrote {len(cases_rows)} case rows')
    print(f'  wrote {len(monthly_totals)} monthly totals')
    if issues:
        print(f'  issues: {len(issues)}')
    print(f'  output -> {out_dir}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
