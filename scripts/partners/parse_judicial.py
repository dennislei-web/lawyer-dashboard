"""
喆律合署律師案件明細 Excel → 乾淨 CSV
-------------------------------------
輸入：指定 folder 中所有 "NNN年XX律師案件明細*.xlsx"
輸出（同輸出資料夾）：
  - profit_share.csv          每律師每月分潤明細（含諮詢/委任/自案各 tier）
  - cases.csv                 每律師每月案件逐筆
  - monthly_totals.csv        每律師每月 喆律分潤 vs 律師分潤 總計
  - coverage.csv              律師 × 月份 覆蓋表（哪些月份有資料）
  - _parse_issues.txt         解析異常、被跳過的工作表
"""
import openpyxl
from pathlib import Path
import csv, os, re, sys, io
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ---------- 設定（可被環境變數覆蓋）----------
_env_input = os.environ.get('PARTNERS_JUDICIAL_INPUT_DIRS', '').strip()
_env_output = os.environ.get('PARTNERS_OUTPUT_DIR', '').strip()
INPUT_DIRS = [d for d in _env_input.split(os.pathsep) if d] if _env_input else [
    r'C:\Users\admin\Desktop\新增資料夾\drive-download-20260418T020634Z-3-001',
    r'C:\Users\admin\Downloads',
]
FILENAME_PATTERN = re.compile(r'(\d{3})年.*?(方心瑜|孫少輔|孫|許致維|劉明潔)律師案件明細')
OUTPUT_DIR = Path(_env_output) if _env_output else Path(r'C:\Users\admin\Desktop\新增資料夾\合署律師分析_output')

LAWYER_ALIASES = {'孫': '孫少輔'}

def normalize_lawyer(name):
    return LAWYER_ALIASES.get(name, name)

def find_input_files():
    results = []
    for d in INPUT_DIRS:
        p = Path(d)
        if not p.exists(): continue
        for f in p.glob('*.xlsx'):
            m = FILENAME_PATTERN.search(f.name)
            if not m: continue
            roc_year = int(m.group(1))
            lawyer = normalize_lawyer(m.group(2))
            results.append((f, roc_year, lawyer))
    return sorted(results, key=lambda x: (x[2], x[1]))

# ---------- sheet 名稱解析 ----------
SHEET_TYPE = {
    'profit': ['分潤'],
    'income': ['收入'],
}

def classify_sheet(sheet_name, default_year):
    """return (type, year_roc, month, note) where type in {'profit','income',None}"""
    name = sheet_name.strip()
    if '副本' in name:
        return (None, None, None, 'duplicate(副本)')
    if name.startswith('工作表') or name in ('工作表1','工作表13'):
        return (None, None, None, 'empty/default sheet')

    is_profit = '分潤' in name
    is_income = '收入' in name
    if not (is_profit or is_income):
        return (None, None, None, 'unclassified')

    # extract year+month digits
    # 形式：11412分潤 / 114年11月收入 / 11月分潤表 / 11212收入明細 / 111501分潤 / 11302分潤表
    digits = re.findall(r'\d+', name)
    year, month = None, None

    # 特殊：111501 = typo 1 + 11501
    for d in digits:
        if len(d) == 6 and d.startswith('1'):
            d2 = d[1:]
            if len(d2) == 5:
                y, mo = int(d2[:3]), int(d2[3:])
                if 100 <= y <= 120 and 1 <= mo <= 12:
                    year, month = y, mo
                    break

    if year is None:
        # try YYYMM (5 digits)
        for d in digits:
            if len(d) == 5:
                y, mo = int(d[:3]), int(d[3:])
                if 100 <= y <= 120 and 1 <= mo <= 12:
                    year, month = y, mo
                    break

    if year is None:
        # try YYYY年MM月 / YYMM
        m = re.search(r'(\d{3})年\D*(\d{1,2})月', name)
        if m:
            year, month = int(m.group(1)), int(m.group(2))
        else:
            # bare month (12月分潤表) — use default year
            m2 = re.search(r'(\d{1,2})月', name)
            if m2:
                mo = int(m2.group(1))
                if 1 <= mo <= 12:
                    year, month = default_year, mo

    if year is None or month is None:
        return (None, None, None, f'cannot parse year/month from "{name}"')

    t = 'profit' if is_profit else 'income'
    return (t, year, month, None)

# ---------- 分潤 sheet 解析 ----------
# 分潤表有幾種區塊：
#   [區塊標題] "一、諮詢委任" / "一、諮詢委任分潤" / "二、XX律師自案" / "二、自案分潤"
#   [諮詢 header] 月份 | 諮詢費（a) | X律師分潤=a*100%
#   [諮詢 data]   YYMM | amount | amount
#   [委任 header] 月份 | 委任費（A） | 引案（B）=A*??% | 咨詢(C)=A*??% | 處理費用(D) | 利潤（E）=... | 喆律分潤=E*XX% | X律師分潤=E*YY%
#   [委任 data]   YYMM | A | B | C | D | E | zhelu | lawyer
#   [自案 header] 月份 | 自案（A） | ...
#   [自案 data]   ditto
#   [合計]       喆律分潤 | total    // 律師分潤 | total

PCT_RE = re.compile(r'\*\s*(-?\d+(?:\.\d+)?)\s*%')

def extract_pct(header_cell):
    if header_cell is None: return None
    m = PCT_RE.search(str(header_cell))
    if m: return float(m.group(1))
    return None

SECTION_HEAD_RE = re.compile(r'^[一二三四五六七八]、')

LAWYER_ONLY_TIERS = {'介紹','追溯','其他'}            # 律師全拿
SPLIT_MISC_TIERS = {'受僱','續委','轉案','合作'}      # 喆律+律師分成（比例因月而異）

def _blank_misc(lawyer, year, month, tier):
    lawyer_only = tier in LAWYER_ONLY_TIERS
    return {
        'lawyer': lawyer, 'year': year, 'month': month, 'tier': tier,
        'commission_A': None, 'refer_pct': None, 'refer_B': None,
        'consult_pct': None, 'consult_C': None, 'proc_D': None,
        'profit_E': None,
        'zhelu_pct': 0.0 if lawyer_only else None,
        'zhelu_amt': 0.0,
        'lawyer_pct': 100.0 if lawyer_only else None,
        'lawyer_amt': 0.0,
        # lawyer-only tiers pre-lock Z so stray grand-total rows don't leak in
        '_z_locked': lawyer_only,
    }

def parse_profit_sheet(ws, lawyer, year, month):
    """Returns (tier_rows, monthly_total)
    monthly_total is DERIVED from tiers below, not from Excel summary rows
    (some Excel files have broken/placeholder total rows).
    """
    tier_rows = []
    current_section = None  # 委任 / 自案 / 介紹 / 追溯 / 合作 / 其他
    misc_tier_row = None    # aggregate row for 介紹/追溯/合作/其他 section

    rows = list(ws.iter_rows(values_only=True))
    i = 0
    while i < len(rows):
        row = rows[i]
        first = str(row[0]).strip() if row[0] is not None else ''
        cells_text = ' '.join(str(v) for v in row if v is not None)

        # "合計：" (with colon) marks end of sections — flush misc and lock out further misc.
        # Bare "合計" is often just a column-sum inside a section (e.g. 介紹佣金 list subtotal).
        if first.startswith('合計') and ('：' in first or ':' in first):
            if misc_tier_row is not None:
                tier_rows.append(misc_tier_row)
                misc_tier_row = None
            current_section = '合計'

        # section markers — when crossing into misc section, open a new misc row
        if SECTION_HEAD_RE.match(first):
            # close previous misc row
            if misc_tier_row is not None:
                tier_rows.append(misc_tier_row)
                misc_tier_row = None
            # "受僱律師自案" / "誠夫律師自案" — co-handled self-case, shared split.
            # Also any 自案 section AFTER the lawyer's own 自案 block = co-handled.
            lawyer_own_self = any(ln in first for ln in ['孫律','許律','劉律','方律'])
            if '受僱' in first or '誠夫' in first:
                current_section = '受僱'
                misc_tier_row = _blank_misc(lawyer, year, month, '受僱')
            elif '自案' in first and not lawyer_own_self and current_section not in (None, '諮詢', '委任'):
                current_section = '受僱'
                misc_tier_row = _blank_misc(lawyer, year, month, '受僱')
            elif '自案' in first:
                current_section = '自案'
            elif '介紹' in first:
                current_section = '介紹'
                misc_tier_row = _blank_misc(lawyer, year, month, '介紹')
            elif '追溯' in first:
                current_section = '追溯'
                misc_tier_row = _blank_misc(lawyer, year, month, '追溯')
            elif '合作' in first:
                current_section = '合作'
                misc_tier_row = _blank_misc(lawyer, year, month, '合作')
            elif '其他' in first or '應付' in first:
                current_section = '其他'
                misc_tier_row = _blank_misc(lawyer, year, month, '其他')
            elif '續委' in first:
                current_section = '續委'
                misc_tier_row = _blank_misc(lawyer, year, month, '續委')
            elif '轉案' in first:
                current_section = '轉案'
                misc_tier_row = _blank_misc(lawyer, year, month, '轉案')
            elif '合計' in first:
                # sheet-wide grand total below — stop attributing anything further
                current_section = '合計'
            else:
                current_section = '委任'

        # consultation header: 月份 | 諮詢費 | ...分潤=a*100%
        if '月份' in cells_text and '諮詢費' in cells_text:
            if i+1 < len(rows):
                data = rows[i+1]
                if data[0] is not None and isinstance(data[1], (int, float)):
                    tier_rows.append({
                        'lawyer': lawyer, 'year': year, 'month': month,
                        'tier': '諮詢',
                        'commission_A': data[1],
                        'refer_pct': None, 'refer_B': None,
                        'consult_pct': None, 'consult_C': None,
                        'proc_D': None,
                        'profit_E': None,
                        'zhelu_pct': 0.0, 'zhelu_amt': 0.0,
                        'lawyer_pct': 100.0, 'lawyer_amt': data[2] if len(data) > 2 else data[1],
                    })
            i += 2
            continue

        # commission/self-case header: 月份 | 委任費 or 自案 | 引案 | 咨詢 | 處理費用 | 利潤 | 喆律分潤 | X律師分潤
        # BUT skip if current_section is a misc tier (合作案件 also has 月份|委任費 header
        # but with different formula — 咨詢=A*10% as lawyer income, no E split)
        in_misc = current_section in ('介紹','追溯','合作','其他')
        if not in_misc and '月份' in cells_text and ('委任費' in cells_text or '自案' in cells_text):
            # determine tier name
            tier_name = current_section if current_section in ('委任','自案','委任2') else (
                '自案' if '自案' in cells_text else '委任')
            # figure column meanings from this header
            headers = [str(v) if v is not None else '' for v in row]
            idx = {}
            for j, h in enumerate(headers):
                if '委任費' in h or '自案' in h: idx['A'] = j
                elif '引案' in h: idx['B'] = j; idx['B_pct'] = extract_pct(h)
                elif '咨詢' in h or '諮詢' in h: idx['C'] = j; idx['C_pct'] = extract_pct(h)
                elif '處理費' in h: idx['D'] = j
                elif '利潤' in h: idx['E'] = j
                elif '喆律分潤' in h: idx['Z'] = j; idx['Z_pct'] = extract_pct(h)
                elif '律師分潤' in h: idx['L'] = j; idx['L_pct'] = extract_pct(h)
            if i+1 < len(rows):
                data = rows[i+1]
                def g(k):
                    j = idx.get(k)
                    return data[j] if j is not None and j < len(data) else None
                # valid data row check
                if g('A') is not None:
                    tier_rows.append({
                        'lawyer': lawyer, 'year': year, 'month': month,
                        'tier': tier_name,
                        'commission_A': g('A'),
                        'refer_pct': idx.get('B_pct'), 'refer_B': g('B'),
                        'consult_pct': idx.get('C_pct'), 'consult_C': g('C'),
                        'proc_D': g('D'),
                        'profit_E': g('E'),
                        'zhelu_pct': idx.get('Z_pct'), 'zhelu_amt': g('Z'),
                        'lawyer_pct': idx.get('L_pct'), 'lawyer_amt': g('L'),
                    })
            i += 2
            continue

        # misc section sub-total rows (介紹/追溯/合作/其他)
        # Only take the FIRST such row per field — subsequent 喆律分潤/X律分潤 rows
        # are usually sub-totals for later sections OR a sheet-wide grand total
        # (which would double-count if added here).
        if misc_tier_row is not None and len(row) > 1 and isinstance(row[1], (int, float)):
            if first == '喆律分潤' and not misc_tier_row.get('_z_locked'):
                misc_tier_row['zhelu_amt'] = row[1]
                misc_tier_row['_z_locked'] = True
            elif first != '喆律分潤' and '律' in first and '分潤' in first and not misc_tier_row.get('_l_locked'):
                misc_tier_row['lawyer_amt'] = row[1]
                misc_tier_row['_l_locked'] = True

        i += 1

    # flush final misc
    if misc_tier_row is not None:
        tier_rows.append(misc_tier_row)

    # drop empty misc rows and internal lock flags
    MISC = LAWYER_ONLY_TIERS | SPLIT_MISC_TIERS
    tier_rows = [r for r in tier_rows
                 if r['tier'] not in MISC
                 or (r['zhelu_amt'] or 0) != 0 or (r['lawyer_amt'] or 0) != 0]
    for r in tier_rows:
        r.pop('_z_locked', None); r.pop('_l_locked', None)

    # derive monthly total from tiers
    def num(x):
        if x is None: return 0.0
        try: return float(x)
        except: return 0.0
    z = l = 0.0
    for r in tier_rows:
        t = r['tier']
        if t == '諮詢':
            l += num(r['lawyer_amt'])
        elif t in ('委任','委任2','自案'):
            z += num(r['refer_B']) + num(r['zhelu_amt'])
            l += num(r['consult_C']) + num(r['lawyer_amt'])
        else:  # 介紹/追溯/合作/其他
            z += num(r['zhelu_amt'])
            l += num(r['lawyer_amt'])
    monthly_total = {'lawyer': lawyer, 'year': year, 'month': month,
                     'zhelu_total': z, 'lawyer_total': l}
    return tier_rows, monthly_total

# ---------- 收入 sheet 解析 ----------
INCOME_COL_MAP = {
    '當事人': 'client',
    '接案人員': 'handlers',
    '金額': 'amount',
    '日期': 'date',
    '備註': 'note',
    '品牌': 'brand',
    '接案所': 'office',
    '部門': 'dept',
    '類型': 'case_type',
    '是否作廢': 'voided',
    '客戶來源': 'source',
    '委任人': 'client',  # alt for 自案 sections
}

SECTION_KEYWORDS = {
    '自案': '自案',
    '介紹': '介紹',
    '承辦': '承辦',
    '原始檔': '承辦',
}

def parse_income_sheet(ws, lawyer, year, month):
    out = []
    current_section = '承辦'  # 持久狀態：only changed by 「XX明細」 header
    idx = None  # header column map
    in_compact_self = False  # 進入 compact 自案段（col1空 col2=client col3=amount）
    inline_override = None  # 本列覆蓋 section（不影響後續 row）

    def _emit_compact(client_val, amount_val, section):
        if client_val is None or not isinstance(amount_val, (int, float)):
            return False
        cs = str(client_val).strip()
        if not cs or cs in ('小計', '合計', '總計', '姓名', '當事人'):
            return False
        out.append({
            'lawyer': lawyer, 'year': year, 'month': month,
            'section': section,
            'client': cs,
            'handlers': None, 'amount': float(amount_val),
            'date': None, 'note': None, 'brand': None,
            'office': None, 'dept': None, 'case_type': None,
            'voided': None, 'source': None,
        })
        return True

    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        if all(v is None for v in vals):
            in_compact_self = False
            continue
        inline_override = None  # reset 每列
        joined = ' '.join(str(v) for v in vals[:13] if v is not None)
        first = str(vals[0]).strip() if vals[0] is not None else ''

        # 1. 「XX明細」header — 持久 section 改變
        for kw, sec in SECTION_KEYWORDS.items():
            if kw in first and '明細' in first:
                current_section = sec
                break

        # 2. 標準段落 marker（first 整格剛好是 "自案"/"介紹" 等）
        if first in SECTION_KEYWORDS and '明細' not in first:
            sec = SECTION_KEYWORDS[first]
            # 判 inline (主表金額位有值) vs compact (col 2-3 有值)
            main_amt_pos = (idx or {}).get('amount', 5)
            main_amt = vals[main_amt_pos] if main_amt_pos < len(vals) else None
            if isinstance(main_amt, (int, float)):
                # Inline 標記：本列覆蓋 section，不影響後續
                inline_override = sec
                # 流程繼續走下方一般 data row 處理
            else:
                # Compact 段開始
                in_compact_self = True
                _emit_compact(vals[1] if len(vals) > 1 else None,
                              vals[2] if len(vals) > 2 else None, sec)
                continue

        # 3. Compact 段續行（col1 空、col2=client、col3=amount）
        if in_compact_self and not first:
            c_client = vals[1] if len(vals) > 1 else None
            c_amt = vals[2] if len(vals) > 2 else None
            if isinstance(c_amt, (int, float)):
                cs = str(c_client or '').strip()
                if cs in ('合計', '小計', '總計'):
                    in_compact_self = False
                    continue
                if c_client is not None and not isinstance(c_client, (int, float, datetime)):
                    if _emit_compact(c_client, c_amt, '自案'):
                        continue
            # 不符合 compact 結構（例：規費 row col2=日期）→ 結束 compact，回到一般處理
            in_compact_self = False

        # Header row detection: contains 當事人/委任人 AND 金額
        if ('當事人' in joined or '委任人' in joined) and '金額' in joined and '日期' in joined:
            idx = {}
            # Only map LEFT table (first occurrence of each column name)
            # Left table ends when we hit right-side summary '合併後' or blank
            right_start = None
            for j, v in enumerate(vals):
                if v is None: continue
                s = str(v).strip()
                if s in ('合併後', '合併', '小計', '合計', '姓名') and 'A' not in idx and j > 3:
                    right_start = j
                    break
            scan_end = right_start if right_start is not None else len(vals)
            for j in range(scan_end):
                v = vals[j]
                if v is None: continue
                s = str(v).strip()
                for kw, field in INCOME_COL_MAP.items():
                    if kw == s and field not in idx:
                        idx[field] = j
                        break
            idx['_scan_end'] = scan_end
            # section hint from first cell
            for kw, sec in SECTION_KEYWORDS.items():
                if kw in first:
                    current_section = sec
            continue

        if idx is None: continue

        # Data row
        def gv(field):
            j = idx.get(field)
            return vals[j] if j is not None and j < len(vals) else None

        client = gv('client')
        if client is None: continue
        client_s = str(client).strip()
        if not client_s or client_s in ('小計','合計','總計','姓名','當事人'):
            continue
        # Reject rows where the 客戶來源 col is None AND amount looks like summary
        amt = gv('amount')
        dt = gv('date')
        # amount 可能是字串；試著正規化
        if isinstance(amt, str):
            try: amt = float(amt.replace(',',''))
            except: pass

        # 日期正規化（保留 datetime 字串）
        date_str = None
        if isinstance(dt, datetime):
            date_str = dt.strftime('%Y-%m-%d')
        elif dt is not None:
            date_str = str(dt)

        out.append({
            'lawyer': lawyer, 'year': year, 'month': month,
            'section': inline_override or current_section,
            'client': client_s,
            'handlers': gv('handlers'),
            'amount': amt,
            'date': date_str,
            'note': gv('note'),
            'brand': gv('brand'),
            'office': gv('office'),
            'dept': gv('dept'),
            'case_type': gv('case_type'),
            'voided': gv('voided'),
            'source': gv('source'),
        })
    return out

# ---------- main ----------
def main():
    files = find_input_files()
    print(f'Found {len(files)} input files:')
    for f, y, l in files:
        print(f'  {y} {l}  -  {f.name}')
    print()

    profit_rows = []
    cases_rows = []
    month_totals = []
    coverage = {}  # (lawyer, year) -> set of months (profit, income)
    issues = []

    for fpath, roc_year, lawyer in files:
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
        except Exception as e:
            issues.append(f'CANNOT OPEN {fpath.name}: {e}')
            continue
        for sheet_name in wb.sheetnames:
            stype, year, month, note = classify_sheet(sheet_name, roc_year)
            if stype is None:
                issues.append(f'{fpath.name} :: {sheet_name}  -  skipped ({note})')
                continue
            if year != roc_year:
                issues.append(f'{fpath.name} :: {sheet_name}  -  year mismatch (filename={roc_year}, parsed={year})')
            ws = wb[sheet_name]
            key = (lawyer, year)
            coverage.setdefault(key, {'profit': set(), 'income': set()})
            if stype == 'profit':
                tiers, total = parse_profit_sheet(ws, lawyer, year, month)
                profit_rows.extend(tiers)
                if total['zhelu_total'] is not None or total['lawyer_total'] is not None:
                    month_totals.append(total)
                if tiers or total['zhelu_total']:
                    coverage[key]['profit'].add(month)
            else:  # income
                rows = parse_income_sheet(ws, lawyer, year, month)
                cases_rows.extend(rows)
                if rows:
                    coverage[key]['income'].add(month)
        wb.close()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # write profit_share.csv
    with open(OUTPUT_DIR/'profit_share.csv', 'w', encoding='utf-8-sig', newline='') as fp:
        w = csv.DictWriter(fp, fieldnames=[
            'lawyer','year','month','tier',
            'commission_A','refer_pct','refer_B','consult_pct','consult_C',
            'proc_D','profit_E',
            'zhelu_pct','zhelu_amt','lawyer_pct','lawyer_amt',
        ])
        w.writeheader()
        for r in profit_rows: w.writerow(r)

    # write cases.csv
    with open(OUTPUT_DIR/'cases.csv', 'w', encoding='utf-8-sig', newline='') as fp:
        w = csv.DictWriter(fp, fieldnames=[
            'lawyer','year','month','section',
            'client','handlers','amount','date','note','brand',
            'office','dept','case_type','voided','source',
        ])
        w.writeheader()
        for r in cases_rows: w.writerow(r)

    # write monthly_totals.csv
    with open(OUTPUT_DIR/'monthly_totals.csv', 'w', encoding='utf-8-sig', newline='') as fp:
        w = csv.DictWriter(fp, fieldnames=['lawyer','year','month','zhelu_total','lawyer_total'])
        w.writeheader()
        for r in month_totals: w.writerow(r)

    # write coverage.csv
    with open(OUTPUT_DIR/'coverage.csv', 'w', encoding='utf-8-sig', newline='') as fp:
        w = csv.writer(fp)
        w.writerow(['lawyer','year','months_profit','months_income','profit_count','income_count'])
        for (lawyer, year), v in sorted(coverage.items()):
            pm = sorted(v['profit']); im = sorted(v['income'])
            w.writerow([lawyer, year,
                        ','.join(f'{m:02d}' for m in pm),
                        ','.join(f'{m:02d}' for m in im),
                        len(pm), len(im)])

    # write issues
    with open(OUTPUT_DIR/'_parse_issues.txt', 'w', encoding='utf-8') as fp:
        fp.write('\n'.join(issues) if issues else 'no issues')

    print(f'\nwrote {len(profit_rows)} profit tier rows')
    print(f'wrote {len(cases_rows)} case rows')
    print(f'wrote {len(month_totals)} monthly totals')
    print(f'coverage: {len(coverage)} lawyer-year pairs')
    print(f'issues: {len(issues)}')
    print(f'\noutput -> {OUTPUT_DIR}')

if __name__ == '__main__':
    main()
