"""Probe 111 年 Excel structure: sheet names, merged cells, column A as dept group."""
from openpyxl import load_workbook

path = r"C:\Users\admin\Downloads\喆律111年度薪資 (1).xlsx"
wb = load_workbook(path, data_only=True)

print(f"All sheets: {wb.sheetnames}")

for sn in ["111年1月", "2月", "12月"]:
    if sn not in wb.sheetnames:
        print(f"\n--- {sn}: 不存在 ---")
        continue
    ws = wb[sn]
    print(f"\n--- {sn} (max_row={ws.max_row}, max_col={ws.max_column}) ---")
    # Header row
    hdr = [ws.cell(row=1, column=c).value for c in range(1, min(ws.max_column+1, 35))]
    print(f"Row 1 header: {hdr}")
    # First 8 rows of column A and B
    print(f"\nColumn A / B / C / D rows 1-15:")
    for r in range(1, 16):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        print(f"  r{r}: A={a!r}  B={b!r}  C={c!r}  D={d!r}")
    # Merged cells in column A
    a_merges = [str(m) for m in ws.merged_cells.ranges if 'A' in str(m).split(':')[0]]
    print(f"\nMerged ranges starting with A col (first 20): {a_merges[:20]}")
    print(f"Total merged ranges: {len(list(ws.merged_cells.ranges))}")
