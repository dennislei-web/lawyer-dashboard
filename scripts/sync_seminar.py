"""講座報名 → 諮詢約成 轉化追蹤 → Supabase ETL（多場講座合併）
來源 Google Sheet「Inbound Leads Pool - 喆律」(客服團隊維護)

抓整本活頁簿 (xlsx 公開匯出)，自動辨識所有「XXXX講座聯繫 / 講座致電」分頁，
每場打上 seminar 標籤後合併寫入 Supabase。新增講座頁(同樣命名)會自動納入。

各講座頁版型大致相同(報名→利衝→聯繫→LINE→約成)，但欄位順序/名稱會有差異：
  - 有些頁沒有「編號」「填表時間」欄
  - 「諮詢需求」有時叫「是否有諮詢需求」
  - 「是否約成」舊頁叫「已約成」
  - 1140930 多一欄「法零已認領」
故用「關鍵字計分」定位表頭列、再以表頭文字對應欄位(容錯)。
無法可靠對應的頁(缺姓名或約成欄)會跳過並印出警告。

每天 09:00 由 GitHub Action 全量 reload。

Usage:
  python sync_seminar.py          # 抓取 → 寫入 Supabase
  python sync_seminar.py --dry    # 只解析印統計，不寫入
"""
from __future__ import annotations
import os, sys, re, json, io
import urllib.request, urllib.parse, urllib.error
from pathlib import Path
from datetime import date, datetime

ENV_PATH = Path(__file__).parent / ".env"
if ENV_PATH.exists():
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

import openpyxl

SHEET_ID = "1LCzLzUifrFNzukV243lriTcBIqMx7a4idJtbDEzdk-M"
XLSX_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
SEM_PATTERN = re.compile(r"(講座聯繫|講座致電|講座.*致電|致電)")

DRY = "--dry" in sys.argv

# 表頭關鍵字(計分定位表頭列用)
HEADER_TOKENS = ["填表時間", "姓名", "Email", "電話", "聯繫管道", "查詢日期",
                 "姓名查詢", "聯繫完成", "利衝建檔", "是否約成", "已約成",
                 "負責同仁", "預約資訊", "諮詢需求"]

# field → 表頭關鍵字(以子字串比對)
FIELD_HINTS = {
    "lead_no":        ["編號"],
    "reg_at":         ["填表時間"],
    "name":           ["填單的姓名", "姓名"],
    "email":          ["Email", "email", "信箱"],
    "phone":          ["電話"],
    "has_need":       ["諮詢需求"],
    "contact_channel":["聯繫管道", "可以透過什麼管道"],
    "line_id":        ["line ID", "line id", "LINE ID"],
    "contact_time":   ["方便聯繫", "方便聯繫您的時間"],
    "help_needed":    ["如何協助"],
    "query_date":     ["查詢日期"],
    "name_query":     ["姓名查詢"],
    "phone_query":    ["電話查詢"],
    "special_status": ["特殊狀況", "喆律註記特殊狀況"],
    "owner":          ["負責同仁"],
    "contact_notes":  ["聯繫狀況摘要"],
    "contacted":      ["聯繫完成"],
    "conflict_filed": ["利衝建檔"],
    "line_url":       ["導入line", "導入LINE"],
    "booked":         ["是否約成", "已約成"],
    "booking_info":   ["預約資訊"],
}


def fetch_workbook():
    req = urllib.request.Request(XLSX_URL, headers={"User-Agent": "Mozilla/5.0"})
    raw = urllib.request.urlopen(req, timeout=180).read()
    return openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)


def norm(c):
    """openpyxl 儲存格 → 標準字串。日期→'YYYY/M/D'、bool→TRUE/FALSE、整數float去.0。"""
    if c is None:
        return ""
    if isinstance(c, bool):
        return "TRUE" if c else "FALSE"
    if isinstance(c, datetime):
        return f"{c.year}/{c.month}/{c.day} {c.hour}:{c.minute}:{c.second}"
    if isinstance(c, date):
        return f"{c.year}/{c.month}/{c.day}"
    if isinstance(c, float):
        return str(int(c)) if c == int(c) else str(c)
    return str(c).strip()


def sheet_matrix(ws, max_rows=400):
    out = []
    for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
        out.append([norm(c) for c in row])
    return out


def truthy(v):
    return str(v).strip().upper() == "TRUE"


def clean(v):
    if not v:
        return ""
    parts = [p.strip().strip(",") for p in re.split(r"[\n]+", v)]
    return " / ".join([p for p in parts if p])


def parse_reg_date(v):
    if not v:
        return None
    m = re.match(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", v.strip())
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3))).isoformat()
    except ValueError:
        return None


def parse_seminar(name):
    """分頁名 → (顯示用講座名, 講座日期)。名稱開頭民國年7碼(1150601)→2026-06-01。"""
    nm = name.strip().strip("「」\"' ")
    sdate = None
    m = re.match(r"(\d{3})(\d{2})(\d{2})", nm)
    if m:
        try:
            sdate = date(int(m.group(1)) + 1911, int(m.group(2)), int(m.group(3))).isoformat()
        except ValueError:
            sdate = None
    return nm, sdate


def find_header(matrix):
    """前 6 列中關鍵字命中最多者為表頭列；需 >=4 命中。"""
    best_i, best_score = None, 0
    for i, r in enumerate(matrix[:6]):
        joined = "".join(r)
        score = sum(1 for t in HEADER_TOKENS if t in joined)
        if score > best_score:
            best_i, best_score = i, score
    return best_i if best_score >= 4 else None


def build_colmap(header):
    cols = {}
    # 先抓排他欄(電話查詢)，避免 phone 誤抓
    for i, h in enumerate(header):
        ht = (h or "").replace("\n", "")
        if "電話查詢" in ht and "phone_query" not in cols:
            cols["phone_query"] = i
    for field, hints in FIELD_HINTS.items():
        if field in cols:
            continue
        for i, h in enumerate(header):
            ht = (h or "").replace("\n", "")
            if i in cols.values():
                continue
            # name / phone 不可抓到「查詢」欄
            if field in ("name", "phone") and "查詢" in ht:
                continue
            if any(hint in ht for hint in hints):
                cols[field] = i
                break
    return cols


def cell(r, i):
    return r[i].strip() if i is not None and i < len(r) and r[i] is not None else ""


def parse_sheet(name, matrix):
    seminar, sdate = parse_seminar(name)
    hi = find_header(matrix)
    if hi is None:
        return None, "找不到表頭列"
    header = matrix[hi]
    cols = build_colmap(header)
    if "name" not in cols or "booked" not in cols:
        return None, f"缺關鍵欄位(name/booked)；偵測到 {sorted(cols)}"

    out, seq = [], 0
    for ridx, r in enumerate(matrix[hi + 1:]):
        g = lambda f: cell(r, cols[f]) if f in cols else ""
        nm_, em, ph, reg = g("name"), g("email"), g("phone"), g("reg_at")
        if not any([nm_, em, ph, reg]):
            continue
        # 過濾掉把表頭/說明文字當資料的列
        if nm_ in ("姓名", "填單的姓名") or "聯繫後" in nm_ and len(nm_) < 6:
            continue
        seq += 1
        lead_no = g("lead_no")
        out.append({
            "seminar": seminar,
            "seminar_date": sdate,
            "lead_key": f"{seminar}#{lead_no or 'seq'+str(seq)}",
            "lead_no": lead_no,
            "row_index": ridx,
            "reg_at": reg,
            "reg_date": parse_reg_date(reg),
            "name": nm_,
            "email": em,
            "phone": ph,
            "has_need": g("has_need") == "是",
            "contact_channel": clean(g("contact_channel")),
            "line_id": g("line_id"),
            "contact_time": clean(g("contact_time")),
            "help_needed": g("help_needed"),
            "query_date": g("query_date"),
            "name_query": g("name_query"),
            "phone_query": g("phone_query"),
            "special_status": g("special_status"),
            "owner": g("owner"),
            "contact_notes": g("contact_notes"),
            "contacted": truthy(g("contacted")),
            "conflict_filed": truthy(g("conflict_filed")),
            "line_url": g("line_url"),
            "booked": truthy(g("booked")),
            "booking_info": g("booking_info"),
        })
    return out, None


def parse_all(wb):
    all_rows, report = [], []
    names = [n for n in wb.sheetnames if SEM_PATTERN.search(n)]
    for name in names:
        rows, err = parse_sheet(name, sheet_matrix(wb[name]))
        if err:
            report.append((name, 0, err))
            continue
        all_rows.extend(rows)
        report.append((name, len(rows), None))
    # lead_key 去重保險
    seen = set()
    for row in all_rows:
        k = row["lead_key"]
        if k in seen:
            k = f"{k}_{row['row_index']}"
            row["lead_key"] = k
        seen.add(k)
    return all_rows, report


def supabase_replace(rows):
    url = os.environ["SUPABASE_URL"]
    key = os.environ["SUPABASE_SERVICE_KEY"]
    headers = {"apikey": key, "Authorization": f"Bearer {key}"}
    req = urllib.request.Request(f"{url}/rest/v1/seminar_leads?lead_key=not.is.null",
                                 method="DELETE",
                                 headers={**headers, "Prefer": "return=minimal"})
    try:
        urllib.request.urlopen(req)
    except urllib.error.HTTPError as e:
        print(f"  [warn] delete: {e.code} {e.read().decode()[:200]}", file=sys.stderr)
    if not rows:
        return 0
    # 分批 upsert(避免單次 payload 過大)
    BATCH = 500
    params = urllib.parse.urlencode({"on_conflict": "lead_key"})
    for i in range(0, len(rows), BATCH):
        chunk = rows[i:i + BATCH]
        data = json.dumps(chunk, ensure_ascii=False, default=str).encode("utf-8")
        req = urllib.request.Request(f"{url}/rest/v1/seminar_leads?{params}", data=data,
                                     method="POST", headers={
            **headers, "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal"})
        try:
            urllib.request.urlopen(req)
        except urllib.error.HTTPError as e:
            raise SystemExit(f"upsert failed: {e.code} {e.read().decode()[:400]}")
    return len(rows)


def print_report(rows, report):
    print(f"\n各講座頁解析結果:")
    for name, n, err in report:
        flag = f"  ⚠ 跳過: {err}" if err else ""
        print(f"  {name[:34]:36} {n:>4} 筆{flag}")
    n = len(rows)
    cf = sum(1 for r in rows if r["conflict_filed"])
    ct = sum(1 for r in rows if r["contacted"] or (r["contact_notes"] or "").strip())
    li = sum(1 for r in rows if r["line_url"])
    bk = sum(1 for r in rows if r["booked"])
    pc = lambda x: f"{x/n*100:.0f}%" if n else "—"
    print(f"\n合計 {n} 筆 ｜ 報名 {n} → 利衝 {cf}({pc(cf)}) → 聯繫處理 {ct}({pc(ct)}) "
          f"→ 導LINE {li}({pc(li)}) → 約成 {bk}({pc(bk)})")


def main():
    wb = fetch_workbook()
    rows, report = parse_all(wb)
    print_report(rows, report)
    if DRY:
        print("\n[dry-run] 未寫入 Supabase")
        return
    n = supabase_replace(rows)
    print(f"\n✓ 已寫入 seminar_leads：{n} 筆")


if __name__ == "__main__":
    main()
