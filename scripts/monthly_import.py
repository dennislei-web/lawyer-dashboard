"""
monthly_import.py
每月諮詢案件匯入腳本（支援 Excel 範本 + 檔案目錄）。

使用方式：
  Mode A: python monthly_import.py --xlsx 202604_cases.xlsx --dir ./case_files/
  Mode B: python monthly_import.py --dir ./case_files/

環境變數（或 .env）：
  SUPABASE_URL=https://xxxxx.supabase.co
  SUPABASE_SERVICE_KEY=eyJxxxxxxxxx
"""

import argparse
import json
import os
import re
import sys
import urllib.request
import urllib.error

# Windows 終端 UTF-8 輸出
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

from pathlib import Path

# 嘗試載入 python-docx
try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False
    print("警告: python-docx 未安裝，將跳過 .docx 檔案的內容讀取")

# 嘗試載入 openpyxl
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 載入 .env
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent / ".env")
except ImportError:
    # 手動載入 .env
    env_path = Path(__file__).parent / ".env"
    if env_path.exists():
        with open(env_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ.setdefault(k.strip(), v.strip())

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates,return=minimal",
}

# 檔名解析正規表達式
# 格式: 律師名_成案/未成案_日期_案件類型(會議記錄/會議紀錄/逐字稿).docx/.txt
# 支援：未成案和日期之間沒有底線、檔名含「的副本」
FILE_PATTERN = re.compile(
    r'^(.+?)_(成案|未成案)_?(\d{8})_?(.+?)\((會議記錄|會議紀錄|逐字稿)\)'
    r'(?:的副本)?'
    r'\.(docx|txt|pdf)$'
)


# ── 檔案讀取 ──

def read_docx(path: str) -> str:
    """讀取 .docx 檔案內容。"""
    if not HAS_DOCX:
        return ""
    try:
        doc = DocxDocument(path)
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception as e:
        print(f"  讀取 docx 失敗: {path} -> {e}")
        return ""


def read_txt(path: str) -> str:
    """讀取 .txt 檔案內容，嘗試多種編碼。"""
    for enc in ["utf-8", "utf-8-sig", "big5", "cp950", "gbk"]:
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except (UnicodeDecodeError, UnicodeError):
            continue
    print(f"  讀取 txt 失敗（編碼問題）: {path}")
    return ""


def read_file_content(path: str) -> str:
    """根據副檔名讀取檔案內容。"""
    ext = Path(path).suffix.lower()
    if ext == ".txt":
        return read_txt(path)
    elif ext == ".docx":
        return read_docx(path)
    elif ext == ".pdf":
        print(f"  跳過 PDF 檔案: {path}")
        return ""
    return ""


# ── 檔名解析 ──

def parse_filename(filename: str):
    """
    解析檔名，回傳 (律師名, is_signed, date_str, case_type, content_type) 或 None。
    content_type: '會議記錄' / '會議紀錄' / '逐字稿'
    """
    # 去除尾部空白和「的副本」等附加文字
    clean_name = filename.strip()
    # 嘗試去掉尾部非副檔名的多餘空格文字 (例如 "xxx.docx 2")
    clean_name = re.sub(r'\.(docx|txt|pdf)\s+.*$', r'.\1', clean_name)

    m = FILE_PATTERN.match(clean_name)
    if not m:
        return None

    lawyer_name = m.group(1).strip()
    is_signed = m.group(2) == "成案"
    date_str = m.group(3)
    case_type = m.group(4).strip()
    content_type = m.group(5)

    return lawyer_name, is_signed, date_str, case_type, content_type


# ── Excel 解析 ──

def parse_xlsx(xlsx_path: str) -> list[dict]:
    """
    解析 Excel 範本，回傳案件清單。
    每筆: {lawyer_name, case_date, case_type, is_signed,
           meeting_record_file, transcript_file, note}
    """
    if not HAS_OPENPYXL:
        print("錯誤: openpyxl 未安裝，無法讀取 xlsx")
        sys.exit(1)

    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["案件清單"] if "案件清單" in wb.sheetnames else wb.active

    cases = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue  # 空行

        lawyer_name = str(row[0]).strip() if row[0] else ""
        case_date = str(row[1]).strip() if row[1] else ""
        case_type = str(row[2]).strip() if row[2] else ""
        signed_str = str(row[3]).strip() if row[3] else ""

        if not lawyer_name or not case_date or not case_type or not signed_str:
            print(f"  第 {row_idx} 行：必填欄位缺失，跳過")
            continue

        # 處理日期格式
        # 可能是 datetime 物件或字串
        if hasattr(row[1], "strftime"):
            case_date = row[1].strftime("%Y-%m-%d")
        elif re.match(r'^\d{8}$', case_date):
            case_date = f"{case_date[:4]}-{case_date[4:6]}-{case_date[6:8]}"
        # 否則假設已經是 YYYY-MM-DD 格式

        is_signed = signed_str == "成案"

        meeting_file = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        transcript_file = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        note = str(row[6]).strip() if len(row) > 6 and row[6] else ""

        cases.append({
            "lawyer_name": lawyer_name,
            "case_date": case_date,
            "case_type": case_type,
            "is_signed": is_signed,
            "meeting_record_file": meeting_file,
            "transcript_file": transcript_file,
            "note": note,
        })

    wb.close()
    return cases


# ── 目錄掃描 ──

def scan_directory(directory: str) -> dict:
    """
    掃描目錄中的檔案，回傳以 (律師名, date_str, case_type) 為 key 的字典。
    value: {is_signed, meeting_record_path, transcript_path}
    """
    files = {}
    dir_path = Path(directory)

    if not dir_path.exists():
        print(f"錯誤: 目錄不存在 -> {directory}")
        sys.exit(1)

    for f in sorted(dir_path.iterdir()):
        if f.is_dir():
            continue
        if f.suffix.lower() not in (".docx", ".txt", ".pdf"):
            continue

        parsed = parse_filename(f.name)
        if not parsed:
            print(f"  跳過（無法解析檔名）: {f.name}")
            continue

        lawyer_name, is_signed, date_str, case_type, content_type = parsed
        formatted_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
        key = (lawyer_name, formatted_date, case_type)

        if key not in files:
            files[key] = {
                "is_signed": is_signed,
                "meeting_record_path": None,
                "transcript_path": None,
            }

        if content_type in ("會議記錄", "會議紀錄"):
            files[key]["meeting_record_path"] = str(f)
        elif content_type == "逐字稿":
            files[key]["transcript_path"] = str(f)

    return files


# ── Supabase API ──

def supabase_request(method: str, endpoint: str, data=None, headers=None):
    """使用 urllib 發送 HTTP 請求。"""
    url = f"{SUPABASE_URL}{endpoint}"
    body = json.dumps(data).encode("utf-8") if data else None
    req = urllib.request.Request(url, data=body, headers=headers or HEADERS, method=method)
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            resp_data = resp.read().decode("utf-8")
            return resp.status, resp_data
    except urllib.error.HTTPError as e:
        resp_data = e.read().decode("utf-8") if e.fp else ""
        return e.code, resp_data


def get_lawyer_map() -> dict:
    """取得律師名稱到 ID 的對應。"""
    hdrs = {**HEADERS, "Prefer": ""}
    status, body = supabase_request(
        "GET",
        "/rest/v1/lawyers?select=id,name&is_active=eq.true",
        headers=hdrs,
    )
    if status != 200:
        print(f"錯誤: 無法取得律師列表 ({status})")
        print(f"  回應: {body[:500]}")
        sys.exit(1)

    lawyers = json.loads(body)
    return {l["name"]: l["id"] for l in lawyers}


# ── 主要邏輯 ──

def build_cases_from_xlsx_and_dir(xlsx_path: str, dir_path: str | None) -> list[dict]:
    """
    Mode A：從 Excel 範本 + 檔案目錄建立案件清單。
    """
    xlsx_cases = parse_xlsx(xlsx_path)
    print(f"從 Excel 讀取到 {len(xlsx_cases)} 筆案件")

    # 掃描目錄中的檔案
    dir_files = {}
    if dir_path:
        dir_files = scan_directory(dir_path)
        print(f"從目錄掃描到 {len(dir_files)} 組檔案")

    result = []
    for case in xlsx_cases:
        meeting_content = ""
        transcript_content = ""

        key = (case["lawyer_name"], case["case_date"], case["case_type"])

        # 優先使用 Excel 中指定的檔名
        if case["meeting_record_file"] and dir_path:
            file_path = Path(dir_path) / case["meeting_record_file"]
            if file_path.exists():
                meeting_content = read_file_content(str(file_path))
            else:
                print(f"  警告: 找不到會議記錄檔 -> {case['meeting_record_file']}")

        if case["transcript_file"] and dir_path:
            file_path = Path(dir_path) / case["transcript_file"]
            if file_path.exists():
                transcript_content = read_file_content(str(file_path))
            else:
                print(f"  警告: 找不到逐字稿檔 -> {case['transcript_file']}")

        # 如果 Excel 沒指定檔名，嘗試從目錄自動比對
        if not meeting_content and key in dir_files and dir_files[key]["meeting_record_path"]:
            meeting_content = read_file_content(dir_files[key]["meeting_record_path"])

        if not transcript_content and key in dir_files and dir_files[key]["transcript_path"]:
            transcript_content = read_file_content(dir_files[key]["transcript_path"])

        result.append({
            "lawyer_name": case["lawyer_name"],
            "case_date": case["case_date"],
            "case_type": case["case_type"],
            "is_signed": case["is_signed"],
            "meeting_record": meeting_content or None,
            "transcript": transcript_content or None,
        })

    return result


def build_cases_from_dir(dir_path: str) -> list[dict]:
    """
    Mode B：只從檔案目錄（檔名解析）建立案件清單。
    """
    dir_files = scan_directory(dir_path)
    print(f"從目錄掃描到 {len(dir_files)} 組檔案")

    result = []
    for (lawyer_name, case_date, case_type), info in dir_files.items():
        meeting_content = ""
        transcript_content = ""

        if info["meeting_record_path"]:
            meeting_content = read_file_content(info["meeting_record_path"])
        if info["transcript_path"]:
            transcript_content = read_file_content(info["transcript_path"])

        result.append({
            "lawyer_name": lawyer_name,
            "case_date": case_date,
            "case_type": case_type,
            "is_signed": info["is_signed"],
            "meeting_record": meeting_content or None,
            "transcript": transcript_content or None,
        })

    return result


def upsert_to_supabase(cases: list[dict], lawyer_map: dict, dry_run: bool = False):
    """匯入案件到 Supabase，回傳統計。"""
    stats = {
        "total": len(cases),
        "success": 0,
        "skipped_no_lawyer": 0,
        "skipped_no_content": 0,
        "errors": 0,
        "skipped_lawyer_names": set(),
    }

    rows = []
    for case in cases:
        lawyer_id = lawyer_map.get(case["lawyer_name"])
        if not lawyer_id:
            stats["skipped_no_lawyer"] += 1
            stats["skipped_lawyer_names"].add(case["lawyer_name"])
            continue

        if not case["meeting_record"] and not case["transcript"]:
            stats["skipped_no_content"] += 1
            continue

        rows.append({
            "lawyer_id": lawyer_id,
            "case_date": case["case_date"],
            "case_type": case["case_type"],
            "is_signed": case["is_signed"],
            "meeting_record": case["meeting_record"],
            "transcript": case["transcript"],
        })

    if dry_run:
        stats["success"] = len(rows)
        return stats, rows

    if not rows:
        return stats, rows

    # 分批 upsert（每批 50 筆）
    batch_size = 50
    for i in range(0, len(rows), batch_size):
        batch = rows[i : i + batch_size]
        status_code, resp_body = supabase_request(
            "POST",
            "/rest/v1/consultation_cases",
            data=batch,
        )
        if status_code in (200, 201):
            stats["success"] += len(batch)
            print(f"  已匯入 {stats['success']}/{len(rows)} 筆")
        else:
            stats["errors"] += len(batch)
            print(f"  匯入失敗 (batch {i // batch_size + 1}): {status_code}")
            print(f"  回應: {resp_body[:500]}")

    return stats, rows


def print_summary(stats: dict, dry_run: bool = False):
    """印出匯入摘要。"""
    prefix = "[Dry Run] " if dry_run else ""
    print()
    print(f"========== {prefix}匯入摘要 ==========")
    print(f"總案件數:           {stats['total']}")
    print(f"成功匯入:           {stats['success']}")
    print(f"跳過(律師不存在):   {stats['skipped_no_lawyer']}")
    print(f"跳過(無內容):       {stats['skipped_no_content']}")
    print(f"錯誤:               {stats['errors']}")
    print("=" * 36)

    if stats["skipped_lawyer_names"]:
        print()
        print("以下律師名稱在資料庫中找不到:")
        for name in sorted(stats["skipped_lawyer_names"]):
            print(f"  - {name}")


def main():
    parser = argparse.ArgumentParser(
        description="每月諮詢案件匯入 Supabase（支援 Excel 範本 + 檔案目錄）"
    )
    parser.add_argument("--xlsx", help="Excel 範本檔案路徑")
    parser.add_argument("--dir", help="案件檔案目錄路徑")
    parser.add_argument("--dry-run", action="store_true", help="只解析，不實際匯入")
    args = parser.parse_args()

    if not args.xlsx and not args.dir:
        parser.error("請至少提供 --xlsx 或 --dir 其中之一")

    # 建立案件清單
    if args.xlsx:
        print(f"讀取 Excel: {args.xlsx}")
        cases = build_cases_from_xlsx_and_dir(args.xlsx, args.dir)
    else:
        print(f"掃描目錄: {args.dir}")
        cases = build_cases_from_dir(args.dir)

    print(f"\n共解析到 {len(cases)} 筆案件")

    if not cases:
        print("沒有可處理的案件")
        return

    # Dry run: 印出解析結果
    if args.dry_run:
        print("\n[Dry Run] 解析結果:")
        for c in cases:
            sign_label = "成案" if c["is_signed"] else "未成案"
            has_mr = "有" if c["meeting_record"] else "無"
            has_ts = "有" if c["transcript"] else "無"
            print(
                f"  {c['lawyer_name']} | {c['case_date']} | {c['case_type']} | "
                f"{sign_label} | 會議記錄:{has_mr} | 逐字稿:{has_ts}"
            )

    # 檢查 Supabase 設定
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("\n錯誤: 請設定 SUPABASE_URL 和 SUPABASE_SERVICE_KEY 環境變數")
        sys.exit(1)

    # 取得律師對應表
    print("\n取得律師對應表...")
    lawyer_map = get_lawyer_map()
    print(f"  找到 {len(lawyer_map)} 位律師")

    # 匯入
    stats, rows = upsert_to_supabase(cases, lawyer_map, dry_run=args.dry_run)
    print_summary(stats, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
