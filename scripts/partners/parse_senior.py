"""
資深律師轉合署 案件明細 Excel → 乾淨 CSV
-----------------------------------------
輸入：drive-download-...-3-001 底下所有 "NNN年XX律師案件明細.xlsx"
  7 位律師：李昭萱、林昀、徐棠娜、許煜婕、陳璽仲、蕭予馨、吳柏慶

輸出：
  - senior_profit_share.csv   每律師每月分潤明細（含 tier / ratio / 當事人）
  - senior_cases.csv          每律師每月案件逐筆（同 cases.csv schema）
  - senior_monthly_totals.csv 每律師每月 喆律分潤 vs 律師分潤 總計
  - _senior_parse_issues.txt  解析異常

Sheet 結構（每月一張，sheet 名 = YYMM 如 11406）：
  1. 頂部：案件明細（XX律師案件 ... 當事人 / 接案人員 / 金額 / 日期 / ...）
  2. 中段：分潤（XX律師分潤 ... 喆律應付 / XX分潤 / 喆律分潤 / XX應付 ...）
  3. 底部：結算表（XX律師提供 ...）— 不解析，僅作資訊

分潤規則（通用）：
  - 比例 0.7 → 諮詢成案：律師 70% / 喆律 30%
  - 比例 0.6 → 喆律轉案：律師 60% / 喆律 40%
  - 比例 0.1 → 自案：律師 90% / 喆律 10%（右表，ratio = 喆律抽成）
  - 比例 1.0 → 諮詢（100% 律師）
  - 比例 0.05 → 成案獎金（諮詢律師才有）
  - 其他比例 → tier='其他'，保留原值
"""
import openpyxl
from pathlib import Path
import csv, os, re, sys, io
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ---------- 設定（可被環境變數覆蓋）----------
_env_input = os.environ.get('PARTNERS_SENIOR_INPUT_DIRS', '').strip()
_env_output = os.environ.get('PARTNERS_OUTPUT_DIR', '').strip()
INPUT_DIRS = [d for d in _env_input.split(os.pathsep) if d] if _env_input else [
    r'C:\Users\admin\Desktop\新增資料夾\drive-download-20260419T022533Z-3-001',
]
SENIOR_LAWYERS = ['李昭萱', '林昀', '徐棠娜', '許煜婕', '陳璽仲', '蕭予馨', '吳柏慶', '柯雪莉']
FILENAME_PATTERN = re.compile(r'(\d{3})年.*?(' + '|'.join(SENIOR_LAWYERS) + r')律師案件明細')
OUTPUT_DIR = Path(_env_output) if _env_output else Path(r'C:\Users\admin\Desktop\新增資料夾\合署律師分析_output')

def find_input_files():
    results = []
    for d in INPUT_DIRS:
        p = Path(d)
        if not p.exists(): continue
        for f in p.glob('*.xlsx'):
            m = FILENAME_PATTERN.search(f.name)
            if not m: continue
            roc_year = int(m.group(1))
            lawyer = m.group(2)
            results.append((f, roc_year, lawyer))
    return sorted(results, key=lambda x: (x[2], x[1]))

# ---------- sheet 分類 ----------
# YYMM 後面允許接任何非數字後綴（容錯「11501收入」「11501分潤」等命名）
YYMM_RE = re.compile(r'^(\d{3})(\d{2})(?!\d)')
MONTH_ONLY_RE = re.compile(r'^(\d{1,2})月')

def classify_sheet(sn, fallback_roc_year=None):
    sn = sn.strip()
    if sn == '綜合': return (None, None, 'summary')
    if sn == '總匯': return (None, None, 'summary')
    # 版型 A：YYMM 例如 11501
    m = YYMM_RE.match(sn)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        if 100 <= y <= 120 and 1 <= mo <= 12:
            return (y, mo, None)
        return (None, None, f'year/month out of range: {sn}')
    # 版型 B：「9月」「10月」— 需要從檔名帶 roc_year
    m = MONTH_ONLY_RE.match(sn)
    if m and fallback_roc_year is not None:
        mo = int(m.group(1))
        if 1 <= mo <= 12:
            return (fallback_roc_year, mo, None)
    return (None, None, f'unrecognised name: {sn}')

# ---------- helpers ----------
def is_num(x):
    return isinstance(x, (int, float)) and not isinstance(x, bool)

def to_num(x, default=0.0):
    if x is None: return default
    if is_num(x): return float(x)
    try: return float(str(x).replace(',', ''))
    except: return default

SECTION_END_KEYWORDS = ('合計', '結算金額', '折抵', '結餘', '備註', '喆律利潤')

def is_row_blank(row, end=12):
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in row[:end])

# ---------- tier 推論 ----------
def tier_from_ratio(side, ratio, tier_hint=None):
    """side = 'left' (喆律端收款，律師抽成) or 'right' (律師端收款，付給喆律)"""
    if tier_hint:
        if '諮詢成案' in tier_hint: return '諮詢成案'
        if '成案獎金' in tier_hint: return '成案獎金'
        if tier_hint == '諮詢': return '諮詢'
        if '轉案' in tier_hint: return '喆律轉案'
        if '自案' in tier_hint: return '自案'
    if ratio is None:
        return '其他'
    r = round(float(ratio), 3)
    if side == 'left':
        if r == 1.0: return '諮詢'
        if r == 0.7: return '諮詢成案'
        if r == 0.6: return '喆律轉案'
        if r == 0.05: return '成案獎金'
        return '其他'
    else:  # right side — ratio = 喆律抽成
        if r == 0.1: return '自案'
        if r in (0.3, 0.35): return '法律010轉案'  # 0.3 原規則, 0.35 後續調整
        return '其他-自案'

# ---------- 案件明細解析 ----------
CASE_COL_MAP = {
    '當事人': 'client',
    '委任人': 'client',
    '接案人員': 'handlers',
    '金額': 'amount',
    '日期': 'date',
    '備註': 'note',
    '品牌': 'brand',
    '接案所': 'office',
    '部門': 'dept',
    '類型': 'case_type',
    '是否作廢': 'voided',
    '客戶來源': 'source',
}

def find_case_header(rows):
    """Return (header_row_idx, col_map) for the case-details table, or None."""
    for i, row in enumerate(rows):
        if i > 5: break  # header is near top
        vals = row
        if not vals: continue
        first = str(vals[0]) if vals[0] is not None else ''
        if '律師案件' not in first and '案件' not in first:
            continue
        # map cols
        idx = {}
        for j, v in enumerate(vals):
            if v is None: continue
            s = str(v).strip()
            if s in CASE_COL_MAP and CASE_COL_MAP[s] not in idx:
                idx[CASE_COL_MAP[s]] = j
        if 'client' in idx and 'amount' in idx:
            return (i, idx)
    return None

def parse_case_section(rows, lawyer, year, month):
    out = []
    found = find_case_header(rows)
    if not found: return out
    header_idx, idx = found
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        if is_row_blank(row): break
        client = row[idx['client']] if idx['client'] < len(row) else None
        if client is None: continue
        client_s = str(client).strip()
        if not client_s or client_s in ('小計', '合計', '總計', '姓名', '當事人'):
            break
        amt = row[idx['amount']] if 'amount' in idx and idx['amount'] < len(row) else None
        if isinstance(amt, str):
            try: amt = float(amt.replace(',', ''))
            except: pass
        dt = row[idx['date']] if 'date' in idx and idx['date'] < len(row) else None
        date_str = dt.strftime('%Y-%m-%d') if isinstance(dt, datetime) else (str(dt) if dt is not None else None)
        def gv(f):
            j = idx.get(f)
            return row[j] if j is not None and j < len(row) else None
        out.append({
            'lawyer': lawyer, 'year': year, 'month': month,
            'section': '承辦',  # senior 律師大多是承辦/轉案；區分交由分潤判斷
            'client': client_s,
            'handlers': gv('handlers'),
            'amount': amt,
            'date': date_str,
            'note': gv('note'),
            'brand': gv('brand'),
            'office': gv('office'),
            'dept': gv('dept'),
            'case_type': gv('case_type'),
            'voided': gv('voided'),
            'source': gv('source'),
        })
    return out

# ---------- 分潤解析 ----------
def find_profit_header(rows):
    """Return list of (idx, right_col) for each profit-table header found.

    支援兩種版型：
      A. senior 版（昭萱/煜婕等）— col 0 = 「XX律師分潤」
      B. consult 版（雪莉等）— col 0 空、col 1 = 「喆律應付」、row 內某 col 含「分潤」
    """
    results = []
    other_markers = ('喆律應付-其他', '喆律應付 - 其他')
    for i, row in enumerate(rows):
        first = str(row[0]).strip() if row[0] is not None else ''
        all_text = ' '.join(str(v) for v in row if v is not None)
        c1s = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
        # 版型 A：col 0 自帶「分潤」
        if '分潤' in first and '喆律應付' in all_text:
            pass  # OK, fall through to right_col detect
        # 版型 B：col 0 空、col 1 = 「喆律應付」、row 含「分潤」或「利潤」
        # （「分潤」雪莉 11501/11502/11504；「利潤」雪莉 11503）
        elif c1s == '喆律應付' and ('分潤' in all_text or '利潤' in all_text):
            pass
        # 版型 C：只有 sub-section（雪莉 114 年），「喆律應付-其他」可能在 col 0 或 col 1
        elif first in other_markers or c1s in other_markers:
            results.append((i, None))  # 沒有 right table
            continue
        else:
            continue
        # find right-table start col (XX應付 != 喆律應付，排除「喆律應付-其他」之類)
        right_col = None
        for j, v in enumerate(row):
            if v is None or j <= 1: continue
            s = str(v).strip()
            if s.endswith('應付') and s != '喆律應付' and '-' not in s:
                right_col = j
                break
        results.append((i, right_col))
    return results

def parse_profit_section(rows, lawyer, year, month):
    """Return list of profit entries."""
    entries = []
    headers = find_profit_header(rows)
    if not headers: return entries

    start_idx, right_col = headers[0]

    # 若 entry point 本身就是「喆律應付-其他」(版型 C，雪莉 114)，直接 in_other 起
    start_row = rows[start_idx]
    start_c0 = str(start_row[0]).strip() if start_row and start_row[0] is not None else ''
    start_c1 = str(start_row[1]).strip() if len(start_row) > 1 and start_row[1] is not None else ''
    other_markers = ('喆律應付-其他', '喆律應付 - 其他')
    in_other_section = start_c0 in other_markers or start_c1 in other_markers

    i = start_idx + 1
    # 「喆律應付-其他」sub-section 的欄位 layout（多種版型）：
    #   senior 版（昭萱等）：client_col=1; c1=client, c2=amount, c3=ratio, c5=tier_text
    #   consult 版 col1（雪莉 115）：client_col=1; c1=client, c2=tier_text, c3=amount, c4=ratio, c5=payable
    #   consult 版 col0（雪莉 114）：client_col=0; col 0=client, 1=tier, 2=amount, 3=ratio, 4=payable
    other_layout = 'senior'   # senior / consult
    other_client_col = 1       # 1 or 0（依 sub-section header 偵測）
    while i < len(rows):
        row = rows[i]
        first = str(row[0]).strip() if row[0] is not None else ''
        c1 = row[1] if len(row) > 1 else None
        c1s = str(c1).strip() if c1 is not None else ''

        # stop if hit end keywords at col 0
        if any(k in first for k in SECTION_END_KEYWORDS):
            break

        # 第二張對帳表（與第一張分潤表資訊重複）的標頭 — 蕭予馨等版型在分潤表
        # 下方還有一個「喆律應付（費用進入的喆律帳戶）/ 客戶 | 案件性質 | 月客戶付款金 ...」
        # 的對帳表，col 1 = 案件性質（如「喆律轉案」），若 senior layout 把它當 client
        # 會產出空當事人的鬼魂 row。在此 break 即可避免。
        if '費用進入' in first or '費用進入' in c1s:
            break
        if c1s == '案件性質' or first == '案件性質':
            break

        # 「喆律應付-其他」標記（col 0 或 col 1）
        if c1s in ('喆律應付-其他', '喆律應付 - 其他') or first in ('喆律應付-其他', '喆律應付 - 其他'):
            in_other_section = True
            i += 1
            continue
        # 偵測 sub-section header「客戶名稱 | 類型 | 實付金額 | 比例 | 應付金額」
        # 「客戶名稱」可能在 col 0（雪莉 114）或 col 1（雪莉 115/senior）
        if in_other_section:
            for off in (1, 0):
                if (len(row) > off + 4 and row[off] is not None
                        and str(row[off]).strip() == '客戶名稱'
                        and '類型' in str(row[off + 1] or '')
                        and '比例' in str(row[off + 3] or '')):
                    other_layout = 'consult'
                    other_client_col = off
                    break
            else:  # no header detected this row
                pass
            if other_layout == 'consult' and (
                (other_client_col == 0 and first == '客戶名稱')
                or (other_client_col == 1 and c1s == '客戶名稱')
            ):
                i += 1
                continue
        # 結尾/摘要列（col 0 或 col 1）
        if c1s in ('姓名', '小計', '合計', '喆律應付', '喆律利潤'):
            i += 1
            continue
        if first in ('小計', '合計', '喆律應付', '喆律利潤'):
            i += 1
            continue

        # 依 layout 取欄位
        if in_other_section and other_layout == 'consult':
            cc = other_client_col
            client_val = row[cc] if cc < len(row) else None
            type_val = row[cc + 1] if cc + 1 < len(row) else None
            amt_cell = row[cc + 2] if cc + 2 < len(row) else None
            ratio_cell = row[cc + 3] if cc + 3 < len(row) else None
            lawyer_amt_cell = row[cc + 4] if cc + 4 < len(row) else None
            client_str = str(client_val).strip() if client_val is not None else ''
            left_tier_hint = type_val.strip() if isinstance(type_val, str) and type_val.strip() in ('諮詢', '成案獎金') else None
            # 退款列可能含「退款」字樣（「巫秀碧-退款」型）
            if not left_tier_hint and isinstance(type_val, str) and '退款' in type_val:
                left_tier_hint = '成案獎金'
        else:
            # senior 版（含主表 row）
            client_str = c1s
            c2 = row[2] if len(row) > 2 else None
            c3 = row[3] if len(row) > 3 else None
            c4 = row[4] if len(row) > 4 else None
            c5 = row[5] if len(row) > 5 else None
            amt_cell, ratio_cell, lawyer_amt_cell = c2, c3, c4
            left_tier_hint = None
            if in_other_section and isinstance(c5, str) and c5.strip() in ('諮詢', '成案獎金'):
                left_tier_hint = c5.strip()

        # left-table data row
        if (client_str and client_str not in ('姓名', '客戶名稱')
                and is_num(amt_cell) and is_num(ratio_cell)
                and 0 < float(ratio_cell) <= 1):
            amt = to_num(amt_cell)
            ratio = float(ratio_cell)
            lawyer_amt = to_num(lawyer_amt_cell) if is_num(lawyer_amt_cell) else amt * ratio
            if in_other_section:
                zhelu_amt = amt - lawyer_amt
            else:
                c5 = row[5] if len(row) > 5 else None
                zhelu_amt = to_num(c5) if is_num(c5) else amt - lawyer_amt
            tier = tier_from_ratio('left', ratio, left_tier_hint)
            entries.append({
                'lawyer': lawyer, 'year': year, 'month': month,
                'side': 'zhelu_handled',
                'tier': tier,
                'client': client_str,
                'case_amount': amt,
                'ratio': ratio,
                'lawyer_amt': lawyer_amt,
                'zhelu_amt': zhelu_amt,
                'note': None,
            })

        # right-table data (律師自案，付給喆律)
        if right_col is not None:
            rc1 = row[right_col] if right_col < len(row) else None
            rc2 = row[right_col + 1] if right_col + 1 < len(row) else None
            rc3 = row[right_col + 2] if right_col + 2 < len(row) else None
            rc1s = str(rc1).strip() if rc1 is not None else ''
            if (rc1s and rc1s not in ('姓名', '小計', '合計')
                and is_num(rc2) and is_num(rc3)
                and 0 < float(rc3) <= 1):
                amt = to_num(rc2)
                ratio = float(rc3)  # ratio = 喆律抽成比例
                zhelu_amt = amt * ratio
                lawyer_amt = amt - zhelu_amt
                tier = tier_from_ratio('right', ratio)
                entries.append({
                    'lawyer': lawyer, 'year': year, 'month': month,
                    'side': 'lawyer_handled',  # 律師收款端
                    'tier': tier,
                    'client': rc1s,
                    'case_amount': amt,
                    'ratio': ratio,
                    'lawyer_amt': lawyer_amt,
                    'zhelu_amt': zhelu_amt,
                    'note': None,
                })

        i += 1

    return entries

# ---------- 備用版型：11504-style「勞務 + 案件報酬分配」 ----------
# 部分律師（如吳柏慶）不用「XX律師分潤 / 喆律應付」結構，改用編號分節：
#   1、勞務                            (左, 喆律端付律師)
#   3、案件報酬分配-直接進入律師帳戶  (右, 律師端收 → 付喆律)
#   4、案件報酬分配-由喆律代收         (右, 喆律端收 → 付律師)
# 當主 parse_profit_section 抓不到分潤段落時，fallback 到本解析器。

# 案件性質字串 → tier。None = 純勞務，不算案件業績、跳過。
CASE_NATURE_TO_TIER = {
    '諮詢': '諮詢',
    '諮詢費': '諮詢',
    '成案獎金': '成案獎金',
    '諮詢後轉回所內': '成案獎金',
    '轉案': '喆律轉案',
    '喆律轉案': '喆律轉案',
    '轉諮': '諮詢成案',  # 30/70 ratio 對應諮詢成案
    '諮詢成案': '諮詢成案',
    '自案': '自案',
    '喆律自案': '自案',
    '法0轉諮': '法律010轉案',
    '法律010轉案': '法律010轉案',
    '代開庭': None,
    '2小時': None,
    '6小時': None,
    '勞務': None,
}

def _cs(row, j):
    if not row or j >= len(row) or row[j] is None: return ''
    return str(row[j]).strip()

def find_alt_tables(rows):
    """偵測 11504-style 表頭。回傳 list[dict]，每個 dict 含 style/header_idx/start_col/direction。"""
    found = []
    for i, row in enumerate(rows):
        if not row: continue
        L = len(row)
        # labor 版型：客戶名稱 | 類型 | 實付金額 | 比例 | 應付金額
        for off in range(0, max(1, min(8, L - 4))):
            if (_cs(row, off) == '客戶名稱'
                    and _cs(row, off + 1) == '類型'
                    and '實付' in _cs(row, off + 2)
                    and '比例' in _cs(row, off + 3)
                    and '應付' in _cs(row, off + 4)):
                found.append({'style': 'labor', 'header_idx': i, 'start_col': off})
                break
        # allocation 版型：編號 | 客戶名稱 | 案件性質 | 給付比例 | 當月客戶付款金額 | 應付給喆律/應付金額
        for off in range(0, max(1, min(10, L - 5))):
            if (_cs(row, off) == '編號'
                    and _cs(row, off + 1) == '客戶名稱'
                    and ('性質' in _cs(row, off + 2) or _cs(row, off + 2) == '類型')
                    and '比例' in _cs(row, off + 3)
                    and ('付款' in _cs(row, off + 4) or '金額' in _cs(row, off + 4))
                    and '應付' in _cs(row, off + 5)):
                last_label = _cs(row, off + 5)
                if '給喆律' in last_label:
                    direction = 'lawyer_collected'
                else:
                    direction = 'zhelu_collected'  # default 給「應付金額」
                    for k in range(max(0, i - 3), i):
                        joined = ' '.join(_cs(rows[k], c) for c in range(len(rows[k])))
                        if '直接進入' in joined:
                            direction = 'lawyer_collected'; break
                        if '喆律代收' in joined:
                            direction = 'zhelu_collected'; break
                found.append({
                    'style': 'allocation', 'header_idx': i,
                    'start_col': off, 'direction': direction,
                })
                break
    return found

def parse_alt_profit_sections(rows, lawyer, year, month):
    """11504-style 備用解析（勞務 + 案件報酬分配）。"""
    out = []
    for h in find_alt_tables(rows):
        style, c0 = h['style'], h['start_col']
        i = h['header_idx'] + 1
        blank_streak = 0
        while i < len(rows):
            row = rows[i]
            if not row:
                blank_streak += 1
                if blank_streak >= 3: break
                i += 1; continue
            first = _cs(row, 0)
            if any(k in first for k in SECTION_END_KEYWORDS): break
            if first.startswith('115年') or first.startswith('116年') or first.startswith('114年'): break

            if style == 'labor':
                client_s = _cs(row, c0)
                type_s = _cs(row, c0 + 1)
                amt_v = row[c0 + 2] if c0 + 2 < len(row) else None
                ratio_v = row[c0 + 3] if c0 + 3 < len(row) else None
                payable_v = row[c0 + 4] if c0 + 4 < len(row) else None
            else:
                client_s = _cs(row, c0 + 1)
                type_s = _cs(row, c0 + 2)
                ratio_v = row[c0 + 3] if c0 + 3 < len(row) else None
                amt_v = row[c0 + 4] if c0 + 4 < len(row) else None
                payable_v = row[c0 + 5] if c0 + 5 < len(row) else None

            if client_s.startswith('小計') or client_s in ('合計', '總計'): break
            if not client_s:
                blank_streak += 1
                if blank_streak >= 3: break
                i += 1; continue
            blank_streak = 0
            if not (is_num(amt_v) and is_num(ratio_v)):
                i += 1; continue

            tier = CASE_NATURE_TO_TIER.get(type_s)
            if tier is None:  # 代開庭/2小時/6小時/勞務 或未知 → skip
                i += 1; continue

            amt = to_num(amt_v)
            ratio = float(ratio_v)
            payable = to_num(payable_v) if is_num(payable_v) else amt * ratio

            if style == 'labor':
                # 喆律端付律師：payable = 律師拿
                lawyer_amt = payable
                zhelu_amt = amt - lawyer_amt
                side = 'zhelu_handled'
                note = '11504-labor'
            elif h['direction'] == 'lawyer_collected':
                # 律師收，付喆律：payable = 喆律拿
                zhelu_amt = payable
                lawyer_amt = amt - zhelu_amt
                side = 'lawyer_handled'
                note = '11504-lawyer-collected'
            else:
                # 喆律代收，付律師：payable = 喆律 keeps
                zhelu_amt = payable
                lawyer_amt = amt - zhelu_amt
                side = 'zhelu_handled'
                note = '11504-zhelu-collected'

            out.append({
                'lawyer': lawyer, 'year': year, 'month': month,
                'side': side, 'tier': tier,
                'client': client_s, 'case_amount': amt, 'ratio': ratio,
                'lawyer_amt': lawyer_amt, 'zhelu_amt': zhelu_amt,
                'note': note,
            })
            i += 1
    return out

# ---------- main ----------
def main():
    files = find_input_files()
    print(f'Found {len(files)} senior-lawyer files:')
    for f, y, l in files:
        print(f'  {y} {l}  -  {f.name}')
    print()

    profit_rows = []
    cases_rows = []
    month_totals = []
    issues = []

    for fpath, roc_year, lawyer in files:
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
        except Exception as e:
            issues.append(f'CANNOT OPEN {fpath.name}: {e}')
            continue

        for sheet_name in wb.sheetnames:
            year, month, note = classify_sheet(sheet_name, fallback_roc_year=roc_year)
            if year is None:
                if note and 'summary' not in note:
                    issues.append(f'{fpath.name} :: {sheet_name}  -  skipped ({note})')
                continue
            if year != roc_year:
                # tolerate — some files bundle 114 and 115 data
                pass

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            # cases
            try:
                cases = parse_case_section(rows, lawyer, year, month)
                cases_rows.extend(cases)
            except Exception as e:
                issues.append(f'{fpath.name} :: {sheet_name}  case-parse failed: {e}')

            # profit
            try:
                entries = parse_profit_section(rows, lawyer, year, month)
                if not entries:
                    # fallback：11504-style「勞務 + 案件報酬分配」版型
                    entries = parse_alt_profit_sections(rows, lawyer, year, month)
                profit_rows.extend(entries)
                # monthly total from entries
                z = sum(e['zhelu_amt'] for e in entries)
                l = sum(e['lawyer_amt'] for e in entries)
                if entries:
                    month_totals.append({
                        'lawyer': lawyer, 'year': year, 'month': month,
                        'zhelu_total': z, 'lawyer_total': l,
                    })
            except Exception as e:
                issues.append(f'{fpath.name} :: {sheet_name}  profit-parse failed: {e}')

        wb.close()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # senior_profit_share.csv
    with open(OUTPUT_DIR / 'senior_profit_share.csv', 'w', encoding='utf-8-sig', newline='') as fp:
        w = csv.DictWriter(fp, fieldnames=[
            'lawyer', 'year', 'month', 'side', 'tier', 'client',
            'case_amount', 'ratio', 'lawyer_amt', 'zhelu_amt', 'note',
        ])
        w.writeheader()
        for r in profit_rows: w.writerow(r)

    # senior_cases.csv — same schema as cases.csv
    with open(OUTPUT_DIR / 'senior_cases.csv', 'w', encoding='utf-8-sig', newline='') as fp:
        w = csv.DictWriter(fp, fieldnames=[
            'lawyer', 'year', 'month', 'section',
            'client', 'handlers', 'amount', 'date', 'note', 'brand',
            'office', 'dept', 'case_type', 'voided', 'source',
        ])
        w.writeheader()
        for r in cases_rows: w.writerow(r)

    # senior_monthly_totals.csv
    with open(OUTPUT_DIR / 'senior_monthly_totals.csv', 'w', encoding='utf-8-sig', newline='') as fp:
        w = csv.DictWriter(fp, fieldnames=['lawyer', 'year', 'month', 'zhelu_total', 'lawyer_total'])
        w.writeheader()
        for r in month_totals: w.writerow(r)

    # issues
    with open(OUTPUT_DIR / '_senior_parse_issues.txt', 'w', encoding='utf-8') as fp:
        fp.write('\n'.join(issues) if issues else 'no issues')

    print(f'\nwrote {len(profit_rows)} profit entries')
    print(f'wrote {len(cases_rows)} case rows')
    print(f'wrote {len(month_totals)} monthly totals')
    print(f'issues: {len(issues)}')
    print(f'\noutput -> {OUTPUT_DIR}')

if __name__ == '__main__':
    main()
